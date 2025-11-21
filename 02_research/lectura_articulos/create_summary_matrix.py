#!/usr/bin/env python3
"""
Script para crear matriz de información resumida del análisis estadístico
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def create_summary_matrix():
    """
    Crea una matriz de información resumida con datos clave
    """
    print("=" * 80)
    print("CREACIÓN DE MATRIZ DE INFORMACIÓN RESUMIDA")
    print("=" * 80)
    
    # Leer datos
    base_dir = Path(__file__).parent
    matriz_path = base_dir / 'matriz_consolidada_v3_20251119_1515.xlsx'
    
    print(f"\n📖 Leyendo datos de: {matriz_path.name}")
    df = pd.read_excel(matriz_path, sheet_name='Datos_Completos')
    print(f"   ✓ {len(df)} artículos cargados")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # ========================================================================
    # HOJA 1: RESUMEN GENERAL
    # ========================================================================
    ws_resumen = wb.active
    ws_resumen.title = 'Resumen General'
    
    # Título
    ws_resumen.merge_cells('A1:D1')
    ws_resumen['A1'] = 'RESUMEN ESTADÍSTICO - BASE DE CONOCIMIENTOS ANH951'
    ws_resumen['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_resumen['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_resumen.row_dimensions[1].height = 30
    
    # Información general
    row = 3
    ws_resumen[f'A{row}'] = 'INFORMACIÓN GENERAL'
    ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    info_general = [
        ['Total de artículos analizados:', len(df)],
        ['Rango temporal:', f"{df['Año'].min():.0f} - {df['Año'].max():.0f}"],
        ['Artículos con autor identificado:', f"{(df['Autores'].notna() & (df['Autores'] != 'No especificado')).sum()} ({(df['Autores'].notna() & (df['Autores'] != 'No especificado')).sum()/len(df)*100:.1f}%)"],
        ['Artículos con hidruro especificado:', f"{(df['Material_Hidruro'].notna() & (df['Material_Hidruro'] != 'No especificado')).sum()} ({(df['Material_Hidruro'].notna() & (df['Material_Hidruro'] != 'No especificado')).sum()/len(df)*100:.1f}%)"],
    ]
    
    for item in info_general:
        ws_resumen[f'A{row}'] = item[0]
        ws_resumen[f'B{row}'] = item[1]
        ws_resumen[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Arquitectura
    row += 1
    ws_resumen[f'A{row}'] = 'ARQUITECTURA DE REACTORES'
    ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    arch_counts = df['Arquitectura'].value_counts()
    for arch, count in arch_counts.items():
        ws_resumen[f'A{row}'] = arch
        ws_resumen[f'B{row}'] = count
        ws_resumen[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # Gestión Térmica
    row += 1
    ws_resumen[f'A{row}'] = 'GESTIÓN TÉRMICA'
    ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    thermal_counts = df['Estrategia_Térmica'].value_counts()
    for thermal, count in thermal_counts.items():
        ws_resumen[f'A{row}'] = thermal
        ws_resumen[f'B{row}'] = count
        ws_resumen[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # Escala
    row += 1
    ws_resumen[f'A{row}'] = 'ESCALABILIDAD'
    ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    # Limpiar escalas largas
    df_clean = df.copy()
    df_clean['Escala_Limpia'] = df_clean['Escala'].apply(
        lambda x: 'Multiescala' if isinstance(x, str) and len(x) > 50 else x
    )
    
    scale_counts = df_clean['Escala_Limpia'].value_counts()
    for scale, count in scale_counts.items():
        ws_resumen[f'A{row}'] = scale
        ws_resumen[f'B{row}'] = count
        ws_resumen[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 50
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    
    # ========================================================================
    # HOJA 2: MATERIALES
    # ========================================================================
    ws_materials = wb.create_sheet('Materiales')
    
    # Título
    ws_materials.merge_cells('A1:D1')
    ws_materials['A1'] = 'MATERIALES DE HIDRURO METÁLICO'
    ws_materials['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_materials['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_materials['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_materials.row_dimensions[1].height = 30
    
    # Encabezados
    headers = ['Material', 'Cantidad', 'Porcentaje', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws_materials.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos de materiales
    material_counts = df['Material_Hidruro'].value_counts()
    row = 4
    for material, count in material_counts.items():
        ws_materials[f'A{row}'] = material
        ws_materials[f'B{row}'] = count
        ws_materials[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        
        # Observaciones según material
        if material == 'TiFe':
            ws_materials[f'D{row}'] = 'Bajo costo, buenas propiedades'
        elif material == 'LaNi5':
            ws_materials[f'D{row}'] = 'Alta capacidad, cinética rápida'
        elif material == 'MgH2':
            ws_materials[f'D{row}'] = 'Alta capacidad gravimétrica'
        elif material == 'AB5':
            ws_materials[f'D{row}'] = 'Aleación tipo A-B5'
        elif material == 'AB2':
            ws_materials[f'D{row}'] = 'Aleación tipo A-B2'
        
        row += 1
    
    ws_materials.column_dimensions['A'].width = 25
    ws_materials.column_dimensions['B'].width = 12
    ws_materials.column_dimensions['C'].width = 15
    ws_materials.column_dimensions['D'].width = 40
    
    # ========================================================================
    # HOJA 3: MÉTODOS TÉRMICOS
    # ========================================================================
    ws_thermal = wb.create_sheet('Métodos Térmicos')
    
    # Título
    ws_thermal.merge_cells('A1:C1')
    ws_thermal['A1'] = 'MÉTODOS DE GESTIÓN TÉRMICA'
    ws_thermal['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_thermal['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_thermal['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_thermal.row_dimensions[1].height = 30
    
    # Métodos Pasivos
    ws_thermal['A3'] = 'MÉTODOS PASIVOS'
    ws_thermal['A3'].font = Font(bold=True, size=12)
    
    headers = ['Método', 'Frecuencia', 'Porcentaje']
    for col, header in enumerate(headers, 1):
        cell = ws_thermal.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Contar métodos pasivos
    pasivos = []
    for val in df['Método_Pasivo'].dropna():
        if val != 'No especificado':
            for metodo in str(val).split(','):
                pasivos.append(metodo.strip())
    
    pasivo_counts = pd.Series(pasivos).value_counts()
    row = 5
    for metodo, count in pasivo_counts.head(10).items():
        ws_thermal[f'A{row}'] = metodo
        ws_thermal[f'B{row}'] = count
        ws_thermal[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # Métodos Activos
    row += 1
    ws_thermal[f'A{row}'] = 'MÉTODOS ACTIVOS'
    ws_thermal[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    for col, header in enumerate(headers, 1):
        cell = ws_thermal.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    row += 1
    
    # Contar métodos activos
    activos = []
    for val in df['Método_Activo'].dropna():
        if val != 'No especificado':
            for metodo in str(val).split(','):
                activos.append(metodo.strip())
    
    activo_counts = pd.Series(activos).value_counts()
    for metodo, count in activo_counts.head(10).items():
        ws_thermal[f'A{row}'] = metodo
        ws_thermal[f'B{row}'] = count
        ws_thermal[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    ws_thermal.column_dimensions['A'].width = 35
    ws_thermal.column_dimensions['B'].width = 15
    ws_thermal.column_dimensions['C'].width = 15
    
    # ========================================================================
    # HOJA 4: TIMELINE
    # ========================================================================
    ws_timeline = wb.create_sheet('Timeline')
    
    # Título
    ws_timeline.merge_cells('A1:C1')
    ws_timeline['A1'] = 'EVOLUCIÓN TEMPORAL DE LA INVESTIGACIÓN'
    ws_timeline['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_timeline['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_timeline['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_timeline.row_dimensions[1].height = 30
    
    # Encabezados
    headers = ['Año', 'Cantidad de Estudios', 'Acumulado']
    for col, header in enumerate(headers, 1):
        cell = ws_timeline.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Datos por año
    year_counts = df['Año'].value_counts().sort_index()
    row = 4
    acumulado = 0
    for year, count in year_counts.items():
        if pd.notna(year):
            acumulado += count
            ws_timeline[f'A{row}'] = int(year)
            ws_timeline[f'B{row}'] = count
            ws_timeline[f'C{row}'] = acumulado
            row += 1
    
    ws_timeline.column_dimensions['A'].width = 15
    ws_timeline.column_dimensions['B'].width = 20
    ws_timeline.column_dimensions['C'].width = 15
    
    # ========================================================================
    # HOJA 5: TOP ESTUDIOS
    # ========================================================================
    ws_top = wb.create_sheet('Top Estudios')
    
    # Título
    ws_top.merge_cells('A1:F1')
    ws_top['A1'] = 'ESTUDIOS MÁS RELEVANTES'
    ws_top['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_top['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_top['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_top.row_dimensions[1].height = 30
    
    # Encabezados
    headers = ['Autores', 'Año', 'Material', 'Arquitectura', 'Escala', 'Mejoras']
    for col, header in enumerate(headers, 1):
        cell = ws_top.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True)
    
    # Top estudios (con mejoras especificadas)
    df_mejoras = df[df['Mejoras_%'].notna() & (df['Mejoras_%'] != 'No especificado')].copy()
    
    row = 4
    for idx, estudio in df_mejoras.head(20).iterrows():
        ws_top[f'A{row}'] = estudio['Autores'] if pd.notna(estudio['Autores']) else 'No especificado'
        ws_top[f'B{row}'] = int(estudio['Año']) if pd.notna(estudio['Año']) else ''
        ws_top[f'C{row}'] = estudio['Material_Hidruro'] if pd.notna(estudio['Material_Hidruro']) else ''
        ws_top[f'D{row}'] = estudio['Arquitectura'] if pd.notna(estudio['Arquitectura']) else ''
        ws_top[f'E{row}'] = estudio['Escala'] if pd.notna(estudio['Escala']) and len(str(estudio['Escala'])) < 50 else 'Ver detalle'
        ws_top[f'F{row}'] = estudio['Mejoras_%'] if pd.notna(estudio['Mejoras_%']) else ''
        
        # Ajustar altura de fila
        ws_top.row_dimensions[row].height = 30
        for col in range(1, 7):
            ws_top.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        
        row += 1
    
    ws_top.column_dimensions['A'].width = 30
    ws_top.column_dimensions['B'].width = 10
    ws_top.column_dimensions['C'].width = 15
    ws_top.column_dimensions['D'].width = 15
    ws_top.column_dimensions['E'].width = 20
    ws_top.column_dimensions['F'].width = 30
    
    # Aplicar bordes a todas las hojas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    # Guardar
    output_path = base_dir / 'Matriz_Informacion_Analisis_Estadistico.xlsx'
    wb.save(output_path)
    
    print(f"\n✅ Matriz de información creada exitosamente!")
    print(f"   📄 Archivo: {output_path.name}")
    print(f"   📊 Hojas creadas: {len(wb.worksheets)}")
    print(f"      1. Resumen General")
    print(f"      2. Materiales")
    print(f"      3. Métodos Térmicos")
    print(f"      4. Timeline")
    print(f"      5. Top Estudios")
    print("=" * 80)


if __name__ == "__main__":
    create_summary_matrix()
