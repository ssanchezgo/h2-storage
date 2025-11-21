import os
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Estilos para el Excel
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Columnas optimizadas para análisis de reactores modulares y gestión térmica
DEFAULT_COLUMNS = [
    "Archivo_Nota",
    "Referencia_PDF",
    "Título_Artículo",
    "Autores",
    "Año",
    "Revista/Fuente",
    "País/Institución",
    "Tipo_Estudio",  # Experimental, Simulación, Review, Diseño
    "Escala",  # Laboratorio, Piloto, Industrial
    "Modularidad",  # Sí/No y descripción
    "Configuración_Reactor",  # Cilíndrico, Multi-tubular, Placas, etc.
    "Diámetro_Reactor_mm",  # Valor numérico separado
    "Longitud_Reactor_mm",  # Valor numérico separado
    "Volumen_Reactor_L",  # Valor numérico separado
    "Dimensiones_Completas",  # Texto descriptivo completo
    "Capacidad_H2_kg",  # Valor numérico
    "Capacidad_H2_wt%",  # Capacidad gravimétrica
    "Capacidad_H2_Completa",  # Texto descriptivo
    "Material_Hidruro",
    "Cantidad_Hidruro_kg",  # Valor numérico
    "Cantidad_Hidruro_Completa",  # Texto descriptivo
    "Sistema_Gestión_Térmica",  # Aletas, tubos, PCM, etc.
    "Tipo_Aletas",  # Anulares, longitudinales, cónicas, disco, etc.
    "Número_Aletas",  # Cantidad de aletas
    "Espesor_Aletas_mm",  # Espesor
    "Espaciado_Aletas_mm",  # Separación entre aletas
    "Número_Tubos",  # Valor numérico para multi-tubulares
    "Diámetro_Tubos_mm",  # Diámetro de tubos internos
    "Intercambiador_Calor",  # Tipo de HX usado
    "Conductividad_Térmica_W_mK",  # Conductividad del lecho
    "Mejora_Conductividad_%",  # Porcentaje de mejora
    "Aditivo_Conductividad",  # ENG, espuma, etc.
    "Porcentaje_Aditivo_%",  # % de aditivo añadido
    "Fluido_Térmico",
    "Flujo_Fluido_L_min",  # Caudal del fluido térmico
    "Temp_Absorción_°C",  # Valor numérico
    "Temp_Desorción_°C",  # Valor numérico
    "Temperatura_Completa",  # Texto descriptivo
    "Presión_Absorción_bar",  # Presión de carga
    "Presión_Desorción_bar",  # Presión de descarga
    "Presión_Completa",  # Texto descriptivo
    "Tiempo_Absorción_min",  # Valor numérico en minutos
    "Tiempo_Desorción_min",  # Valor numérico en minutos
    "Mejora_Tiempo_%",  # Mejora porcentual vs diseño base
    "Tiempos_Completos",  # Texto descriptivo
    "Eficiencia_Sistema_%",  # Eficiencia reportada
    "Resultados_Clave",
    "Ventajas_Diseño",
    "Limitaciones",
    "Conclusiones_Modularidad",
    "Conclusiones_Gestión_Térmica",
    "Aplicabilidad_Proyecto",
    "Imágenes_Disponibles"
]


def clean_value(text):
    """Limpia y normaliza el texto extraído."""
    if not text:
        return ""
    # Eliminar múltiples espacios
    text = re.sub(r"\s+", " ", text)
    # Eliminar marcadores markdown extremos pero mantener el formato importante
    text = re.sub(r"\*\*\*\*(.*?)\*\*\*\*", r"\1", text)
    text = re.sub(r"\*\*\*(.*?)\*\*\*", r"\1", text)
    # Limpiar espacios al inicio y final
    return text.strip()


def convert_to_mm(value_str):
    """Convierte dimensiones a milímetros."""
    if not value_str:
        return ""
    
    # Buscar número y unidad
    match = re.search(r'([\d.,]+)\s*(mm|cm|m)\b', value_str, re.IGNORECASE)
    if not match:
        return ""
    
    try:
        number = float(match.group(1).replace(',', '.'))
        unit = match.group(2).lower()
        
        # Convertir a mm
        if unit == 'cm':
            return str(round(number * 10, 2))
        elif unit == 'm':
            return str(round(number * 1000, 2))
        else:  # ya está en mm
            return str(round(number, 2))
    except:
        return ""


def convert_to_liters(value_str):
    """Convierte volumen a litros."""
    if not value_str:
        return ""
    
    match = re.search(r'([\d.,]+)\s*(L|m³|cm³|ml)\b', value_str, re.IGNORECASE)
    if not match:
        return ""
    
    try:
        number = float(match.group(1).replace(',', '.'))
        unit = match.group(2).lower()
        
        # Convertir a litros
        if unit == 'm³' or unit == 'm3':
            return str(round(number * 1000, 2))
        elif unit == 'cm³' or unit == 'cm3':
            return str(round(number / 1000, 2))
        elif unit == 'ml':
            return str(round(number / 1000, 2))
        else:  # ya está en L
            return str(round(number, 2))
    except:
        return ""


def convert_to_kg(value_str):
    """Convierte masa a kilogramos."""
    if not value_str:
        return ""
    
    match = re.search(r'([\d.,]+)\s*(kg|g)\b', value_str, re.IGNORECASE)
    if not match:
        return ""
    
    try:
        number = float(match.group(1).replace(',', '.'))
        unit = match.group(2).lower()
        
        # Convertir a kg
        if unit == 'g':
            return str(round(number / 1000, 4))
        else:  # ya está en kg
            return str(round(number, 4))
    except:
        return ""


def convert_to_celsius(value_str):
    """Convierte temperatura a Celsius y extrae valores."""
    if not value_str:
        return "", ""
    
    # Buscar rangos de temperatura
    range_match = re.search(r'([\d.,]+)\s*[-–~]\s*([\d.,]+)\s*[°]?([CK])', value_str, re.IGNORECASE)
    if range_match:
        try:
            temp1 = float(range_match.group(1).replace(',', '.'))
            temp2 = float(range_match.group(2).replace(',', '.'))
            unit = range_match.group(3).upper()
            
            # Convertir de Kelvin si es necesario
            if unit == 'K':
                temp1 -= 273.15
                temp2 -= 273.15
            
            return str(round(temp1, 1)), str(round(temp2, 1))
        except:
            pass
    
    # Buscar temperatura simple
    simple_match = re.search(r'([\d.,]+)\s*[°]?([CK])', value_str, re.IGNORECASE)
    if simple_match:
        try:
            temp = float(simple_match.group(1).replace(',', '.'))
            unit = simple_match.group(2).upper()
            
            if unit == 'K':
                temp -= 273.15
            
            return str(round(temp, 1)), ""
        except:
            pass
    
    return "", ""


def convert_to_bar(value_str):
    """Convierte presión a bar."""
    if not value_str:
        return ""
    
    match = re.search(r'([\d.,]+)\s*(bar|MPa|atm)\b', value_str, re.IGNORECASE)
    if not match:
        return ""
    
    try:
        number = float(match.group(1).replace(',', '.'))
        unit = match.group(2).lower()
        
        # Convertir a bar
        if unit == 'mpa':
            return str(round(number * 10, 2))
        elif unit == 'atm':
            return str(round(number * 1.01325, 2))
        else:  # ya está en bar
            return str(round(number, 2))
    except:
        return ""


def convert_to_minutes(value_str):
    """Convierte tiempo a minutos."""
    if not value_str:
        return ""
    
    match = re.search(r'([\d.,]+)\s*(s|min|h)\b', value_str, re.IGNORECASE)
    if not match:
        return ""
    
    try:
        number = float(match.group(1).replace(',', '.'))
        unit = match.group(2).lower()
        
        # Convertir a minutos
        if unit == 's':
            return str(round(number / 60, 2))
        elif unit == 'h':
            return str(round(number * 60, 2))
        else:  # ya está en min
            return str(round(number, 2))
    except:
        return ""


def extract_number_of_tubes(content):
    """Extrae el número de tubos en reactores multi-tubulares."""
    patterns = [
        r'(\d+)\s*tubos',
        r'(\d+)[- ]tube',
        r'multi[- ]?tubular[^:]*:\s*(\d+)',
        r'(\d+)[- ]tubular'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            return match.group(1)
    
    return ""


def extract_percentage(text):
    """Extrae valores porcentuales."""
    if not text:
        return ""
    
    match = re.search(r'([\d.,]+)\s*%', text)
    if match:
        try:
            return str(round(float(match.group(1).replace(',', '.')), 2))
        except:
            return ""
    return ""


def extract_conductivity(content):
    """Extrae valores de conductividad térmica."""
    patterns = [
        r'conductividad[^:]{0,30}:\s*([\d.,]+)\s*(?:W/m[·.]?K|W/mK)',
        r'([\d.,]+)\s*W/m[·.]?K',
        r'thermal\s+conductivity[^:]{0,30}:\s*([\d.,]+)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            try:
                return str(round(float(match.group(1).replace(',', '.')), 2))
            except:
                pass
    return ""


def extract_fin_dimensions(content):
    """Extrae dimensiones y cantidad de aletas."""
    num_aletas = ""
    espesor = ""
    espaciado = ""
    
    # Número de aletas (incluyendo rangos)
    num_patterns = [
        r'(\d+)\s*aletas',
        r'(\d+)\s*fins',
        r'(\d+)[-–]\s*(\d+)\s*aletas',  # Rangos como "8-12 aletas"
        r'aletas[^:]{0,20}:\s*(\d+)',
        r'n[úu]mero[^:]{0,10}aletas[^:]{0,10}:\s*(\d+)'
    ]
    for pattern in num_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            groups = [g for g in match.groups() if g]
            if len(groups) == 2 and groups[1].isdigit():
                # Es un rango, tomar el promedio
                num1, num2 = int(groups[0]), int(groups[1])
                num_aletas = str((num1 + num2) / 2)
            else:
                num_aletas = groups[0] if groups else match.group(1)
            break
    
    # Espesor de aletas (patrones más amplios)
    esp_patterns = [
        r'(\d+)\s*aletas[^:]{0,30}?(?:de\s+)?([\d.,]+)\s*mm\s+(?:de\s+)?espesor',  # "30 aletas de 1.5mm espesor"
        r'([\d.,]+)\s*mm\s+(?:de\s+)?espesor',
        r'espesor[^:]{0,30}?([\d.,]+)\s*mm',
        r'aletas?[^:]{0,50}?([\d.,]+)\s*mm\s+espesor',
        r'fin\s+thickness[^:]{0,10}:\s*([\d.,]+)\s*(?:mm|cm)',
        r'grosor[^:]{0,30}?([\d.,]+)\s*mm'
    ]
    for pattern in esp_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            # Si hay dos grupos capturados, el segundo es el espesor
            all_groups = match.groups()
            if len(all_groups) >= 2 and all_groups[1] is not None:
                espesor = convert_to_mm(all_groups[1])
            elif len(all_groups) >= 1 and all_groups[0] is not None:
                espesor = convert_to_mm(all_groups[0])
            break
    
    # Espaciado entre aletas (patrones más amplios)
    espac_patterns = [
        r'espaciado\s+([\d.,]+)\s*mm',
        r'separaci[óo]n[^:]{0,30}?([\d.,]+)\s*mm',
        r'distancia[^:]{0,30}aletas?[^:]{0,30}?([\d.,]+)\s*mm',
        r'fin\s+spacing[^:]{0,10}:\s*([\d.,]+)\s*(?:mm|cm)',
        r'pitch[^:]{0,10}:\s*([\d.,]+)\s*(?:mm|cm)'
    ]
    for pattern in espac_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            espaciado = convert_to_mm(match.group(1))
            break
    
    return num_aletas, espesor, espaciado


def extract_tube_diameter(content):
    """Extrae el diámetro de tubos internos."""
    patterns = [
        r'diámetro[^:]{0,20}(?:del\s+)?tubo[^:]{0,10}:\s*([\d.,]+)\s*(?:mm|cm)',
        r'tubo[^:]{0,20}diámetro[^:]{0,10}:\s*([\d.,]+)\s*(?:mm|cm)',
        r'tubos?\s+(?:agua|HTF)?[^:]{0,20}?diámetro\s+([\d.,]+)(?:[-–]\s*([\d.,]+))?\s*(?:mm|cm)',
        r'tubos?\s+de\s+([\d.,]+)(?:[-–]\s*([\d.,]+))?\s*(?:mm|cm)',
        r'tube\s+diameter[^:]{0,10}:\s*([\d.,]+)\s*(?:mm|cm)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            groups = [g for g in match.groups() if g]
            if len(groups) == 2:
                # Es un rango, tomar el promedio
                val1 = float(groups[0].replace(',', '.'))
                val2 = float(groups[1].replace(',', '.'))
                avg = (val1 + val2) / 2
                # Convertir a mm si está en cm
                if 'cm' in pattern.lower():
                    avg *= 10
                return str(avg)
            else:
                return convert_to_mm(groups[0] if groups else match.group(1))
    
    return ""


def extract_additive_info(content):
    """Extrae información sobre aditivos para mejorar conductividad."""
    aditivo = ""
    porcentaje = ""
    
    # Tipo de aditivo
    aditivos = {
        'ENG': r'\bENG\b|grafito\s+expandido|expanded\s+(?:natural\s+)?graphite',
        'Espuma metálica': r'espuma\s+metálica|metal\s+foam|copper\s+foam|aluminum\s+foam',
        'Grafito': r'\bgrafito\b|graphite',
        'Compuesto': r'compuesto|composite|MH[- ]composite'
    }
    
    for nombre, pattern in aditivos.items():
        if re.search(pattern, content, re.IGNORECASE):
            aditivo = nombre
            break
    
    # Porcentaje de aditivo
    if aditivo:
        pct_patterns = [
            rf'{aditivo}[^:{{0,30}}]:\s*([\d.,]+)\s*(?:%|wt%)',
            r'([\d.,]+)\s*(?:%|wt%)[^:]{0,20}(?:ENG|grafito|foam)',
            r'([\d.,]+)\s*wt%'
        ]
        for pattern in pct_patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                try:
                    porcentaje = str(round(float(match.group(1).replace(',', '.')), 2))
                    break
                except:
                    pass
    
    return aditivo, porcentaje


def extract_flow_rate(content):
    """Extrae el caudal del fluido térmico."""
    patterns = [
        r'flujo[^:]{0,20}:\s*([\d.,]+)\s*(?:L/min|l/min|lpm)',
        r'caudal[^:]{0,20}:\s*([\d.,]+)\s*(?:L/min|l/min|lpm)',
        r'flow\s+rate[^:]{0,10}:\s*([\d.,]+)\s*(?:L/min|l/min|lpm)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            try:
                return str(round(float(match.group(1).replace(',', '.')), 2))
            except:
                pass
    return ""


def extract_improvement_percentage(content):
    """Extrae porcentajes de mejora en rendimiento."""
    patterns = [
        r'([\d.,]+)\s*%\s+(?:más\s+rápido|faster|mejora|improvement)',
        r'(?:mejora|improvement|reducción|reduction)[^:]{0,30}:\s*([\d.,]+)\s*%',
        r'([\d.,]+)\s*%\s+(?:mejor|better)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            try:
                return str(round(float(match.group(1).replace(',', '.')), 2))
            except:
                pass
    return ""


def extract_efficiency(content):
    """Extrae eficiencia del sistema."""
    patterns = [
        r'eficiencia[^:]{0,20}:\s*([\d.,]+)\s*%',
        r'efficiency[^:]{0,20}:\s*([\d.,]+)\s*%',
        r'([\d.,]+)\s*%\s+(?:de\s+)?eficiencia'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            try:
                return str(round(float(match.group(1).replace(',', '.')), 2))
            except:
                pass
    return ""


def extract_gravimetric_capacity(content):
    """Extrae capacidad gravimétrica en wt%."""
    patterns = [
        r'([\d.,]+)\s*wt%',
        r'([\d.,]+)\s*%\s*(?:en\s+)?peso',
        r'capacidad\s+gravimétrica[^:]{0,20}:\s*([\d.,]+)\s*%'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            try:
                return str(round(float(match.group(1).replace(',', '.')), 2))
            except:
                pass
    return ""


def extract_reference_pdf(content):
    """Extrae la referencia al PDF original."""
    pattern = r"\*\*Referencia PDF:\*\*\s*`([^`]+)`"
    match = re.search(pattern, content, re.IGNORECASE)
    return match.group(1) if match else ""


def extract_year(content, filename):
    """Extrae el año del artículo."""
    # Primero intentar desde el nombre del archivo
    year_match = re.search(r"(19|20)\d{2}", filename)
    if year_match:
        return year_match.group(0)
    
    # Patrones en el contenido
    patterns = [
        r"\((\d{4})\)",  # (2024)
        r"year\s*=\s*\{(\d{4})\}",
        r"\*\*Año:\*\*\s*(\d{4})",
        r"(?:19|20)\d{2}"
    ]
    
    for pattern in patterns:
        match = re.search(pattern, content)
        if match:
            return match.group(1) if match.lastindex else match.group(0)
    
    return ""


def extract_modular_info(content):
    """Extrae información específica sobre modularidad."""
    modular_keywords = [
        r"modular(?:idad)?",
        r"módulo[s]?",
        r"escalab(?:le|ilidad)",
        r"apilable",
        r"stack",
        r"multi(?:-|\s)?etapa",
        r"cascada",
        r"replicable"
    ]
    
    modular_info = []
    for keyword in modular_keywords:
        matches = re.finditer(keyword, content, re.IGNORECASE)
        for match in matches:
            # Extraer contexto (200 caracteres alrededor)
            start = max(0, match.start() - 100)
            end = min(len(content), match.end() + 100)
            context = content[start:end].replace("\n", " ")
            modular_info.append(clean_value(context))
    
    return " | ".join(modular_info[:3]) if modular_info else ""


def extract_thermal_management(content):
    """Extrae información sobre gestión térmica."""
    thermal_keywords = [
        r"gestión\s+térmica",
        r"transferencia\s+de\s+calor",
        r"intercambiador(?:\s+de\s+calor)?",
        r"aletas?",
        r"fins?",
        r"tubos?\s+(?:de\s+)?calor",
        r"heat\s+pipe",
        r"PCM",
        r"material(?:es)?\s+de\s+cambio\s+de\s+fase"
    ]
    
    thermal_info = []
    for keyword in thermal_keywords:
        matches = re.finditer(keyword, content, re.IGNORECASE)
        for match in matches:
            start = max(0, match.start() - 100)
            end = min(len(content), match.end() + 100)
            context = content[start:end].replace("\n", " ")
            thermal_info.append(clean_value(context))
    
    return " | ".join(thermal_info[:3]) if thermal_info else ""


def extract_key_results(content):
    """Extrae resultados clave del estudio."""
    sections = [
        r"Resultados(?:\s+Clave)?(?:\s+Experimentales)?[:\n]+(.*?)(?=\n##|\n---|\Z)",
        r"Conclusiones[:\n]+(.*?)(?=\n##|\n---|\Z)",
        r"Hallazgos(?:\s+Principales)?[:\n]+(.*?)(?=\n##|\n---|\Z)"
    ]
    
    for pattern in sections:
        match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
        if match:
            result = match.group(1)
            # Limitar a primeras 500 caracteres
            return clean_value(result[:500]) + ("..." if len(result) > 500 else "")
    
    return ""


def extract_specific_data(file_path):
    """Extrae información específica del archivo markdown."""
    with open(file_path, "r", encoding="utf-8") as file:
        content = file.read()
    
    filename = Path(file_path).stem
    
    # Inicializar datos
    extracted_data = {col: "" for col in DEFAULT_COLUMNS}
    extracted_data["Archivo_Nota"] = filename
    extracted_data["Referencia_PDF"] = extract_reference_pdf(content)
    extracted_data["Año"] = extract_year(content, filename)
    
    # Patrones principales
    patterns = {
        "Título_Artículo": [
            r"# Notas sobre \"(.*?)\"",
            r"# (.*?)(?:\n|$)",
            r"title\s*=\s*\{([^}]+)\}"
        ],
        "Autores": [
            r"\*\*Autor(?:es)?:\*\*\s*(.*?)(?=\n|$)",
            r"author\s*=\s*\{([^}]+)\}",
            r"Por:\s*(.*?)(?=\n|$)"
        ],
        "Revista/Fuente": [
            r"\*\*(?:Revista|Journal|Fuente):\*\*\s*(.*?)(?=\n|$)",
            r"journal\s*=\s*\{([^}]+)\}"
        ],
        "País/Institución": [
            r"\*\*Institución:\*\*\s*(.*?)(?=\n|$)",
            r"\*\*País:\*\*\s*(.*?)(?=\n|$)",
            r"\*\*Universidad:\*\*\s*(.*?)(?=\n|$)"
        ],
        "Tipo_Estudio": [
            r"\*\*Tipo(?:\s+de\s+estudio)?:\*\*\s*(.*?)(?=\n|$)",
            r"(?:Estudio\s+)?(Experimental|Simulación|CFD|Review|Diseño|Teórico)"
        ],
        "Escala": [
            r"\*\*Escala:\*\*\s*(.*?)(?=\n|$)",
            r"(?:escala\s+)(laboratorio|piloto|industrial|banco|bench)"
        ],
        "Configuración_Reactor": [
            r"\*\*Configuración:\*\*\s*(.*?)(?=\n|$)",
            r"\*\*(?:Tipo\s+de\s+)?(?:reactor|tanque):\*\*\s*(.*?)(?=\n|$)",
            r"(?:reactor|tanque)\s+(?:de\s+tipo\s+)?(cilíndrico|tubular|multi(?:-|\s)?tubular|placas|anular)"
        ],
        "Material_Hidruro": [
            r"\*\*(?:Material|Hidruro):\*\*\s*(.*?)(?=\n|$)",
            r"(LaNi5|LaNi₅|MgH2|MgH₂|TiFe|NaAlH4|NaAlH₄|AB5|AB2)"
        ],
        "Tipo_Aletas": [
            r"aletas?\s+(anulares?|longitudinales?|radiales?|cónicas?|disco|helicoidales?|honeycomb)",
            r"\*\*Aletas:\*\*\s*(.*?)(?=\n|$)"
        ],
        "Fluido_Térmico": [
            r"\*\*Fluido(?:\s+térmico)?:\*\*\s*(.*?)(?=\n|$)",
            r"fluido\s+(?:de\s+)?(agua|aceite|aire|térmico)"
        ]
    }
    
    # Extracción con patrones
    for key, pattern_list in patterns.items():
        for pattern in pattern_list:
            match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
            if match:
                extracted_data[key] = clean_value(match.group(1))
                break
    
    # Patrones numéricos específicos con extracción de texto completo
    numeric_patterns = {
        "Dimensiones_Completas": r"(?:dimensiones?|medidas?):\s*([^\n]{10,100})",
        "Capacidad_H2_Completa": r"capacidad[^:]{0,20}:\s*([^\n]{5,100})",
        "Cantidad_Hidruro_Completa": r"(?:cantidad|masa)[^:]{0,20}:\s*([^\n]{5,100})",
        "Temperatura_Completa": r"temperatura[^:]{0,30}:\s*([^\n]{5,100})",
        "Presión_Completa": r"presión[^:]{0,30}:\s*([^\n]{5,100})",
        "Tiempos_Completos": r"(?:tiempo|absorción|desorción)[^:]{0,30}:\s*([^\n]{5,100})"
    }
    
    for key, pattern in numeric_patterns.items():
        if not extracted_data[key]:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                extracted_data[key] = clean_value(match.group(1))
    
    # Extraer y convertir dimensiones específicas del reactor
    # Diámetro del reactor
    diam_patterns = [
        r'\*\*Diámetro(?:\s+del\s+reactor)?:\*\*\s*([^\n]+)',
        r'diámetro[^:]{0,20}reactor[^:]{0,10}:\s*([^\n]+)',
        r'D\s*=\s*([^\n]+)'
    ]
    for pattern in diam_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            extracted_data["Diámetro_Reactor_mm"] = convert_to_mm(match.group(1))
            break
    
    # Longitud del reactor
    long_patterns = [
        r'\*\*Longitud(?:\s+del\s+reactor)?:\*\*\s*([^\n]+)',
        r'longitud[^:]{0,20}reactor[^:]{0,10}:\s*([^\n]+)',
        r'L\s*=\s*([^\n]+)'
    ]
    for pattern in long_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            extracted_data["Longitud_Reactor_mm"] = convert_to_mm(match.group(1))
            break
    
    # Volumen del reactor
    vol_patterns = [
        r'\*\*Volumen(?:\s+del\s+reactor)?:\*\*\s*([^\n]+)',
        r'volumen[^:]{0,20}reactor[^:]{0,10}:\s*([^\n]+)',
        r'V\s*=\s*([^\n]+)'
    ]
    for pattern in vol_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            extracted_data["Volumen_Reactor_L"] = convert_to_liters(match.group(1))
            break
    
    # Capacidad H2
    cap_match = re.search(r'capacidad[^:]{0,20}H[₂2][^:]{0,20}:\s*([^\n]+)', content, re.IGNORECASE)
    if cap_match:
        extracted_data["Capacidad_H2_kg"] = convert_to_kg(cap_match.group(1))
    
    # Capacidad gravimétrica
    extracted_data["Capacidad_H2_wt%"] = extract_gravimetric_capacity(content)
    
    # Cantidad de hidruro
    cant_match = re.search(r'(?:cantidad|masa)[^:]{0,20}hidruro[^:]{0,20}:\s*([^\n]+)', content, re.IGNORECASE)
    if cant_match:
        extracted_data["Cantidad_Hidruro_kg"] = convert_to_kg(cant_match.group(1))
    elif extracted_data["Cantidad_Hidruro_Completa"]:
        extracted_data["Cantidad_Hidruro_kg"] = convert_to_kg(extracted_data["Cantidad_Hidruro_Completa"])
    
    # ===== EXTRACCIÓN DE PARÁMETROS DE GESTIÓN TÉRMICA =====
    
    # Información de aletas
    num_aletas, espesor_aletas, espaciado_aletas = extract_fin_dimensions(content)
    extracted_data["Número_Aletas"] = num_aletas
    extracted_data["Espesor_Aletas_mm"] = espesor_aletas
    extracted_data["Espaciado_Aletas_mm"] = espaciado_aletas
    
    # Diámetro de tubos internos
    extracted_data["Diámetro_Tubos_mm"] = extract_tube_diameter(content)
    
    # Conductividad térmica
    extracted_data["Conductividad_Térmica_W_mK"] = extract_conductivity(content)
    
    # Mejora en conductividad
    if "mejora" in content.lower() and "conductividad" in content.lower():
        mejora_cond_match = re.search(r'(?:mejora|aumento)[^:]{0,30}conductividad[^:]{0,30}:\s*([\d.,]+)\s*%', content, re.IGNORECASE)
        if mejora_cond_match:
            extracted_data["Mejora_Conductividad_%"] = extract_percentage(mejora_cond_match.group(0))
    
    # Aditivos para mejorar conductividad
    aditivo, pct_aditivo = extract_additive_info(content)
    extracted_data["Aditivo_Conductividad"] = aditivo
    extracted_data["Porcentaje_Aditivo_%"] = pct_aditivo
    
    # Flujo del fluido térmico
    extracted_data["Flujo_Fluido_L_min"] = extract_flow_rate(content)
    
    # Mejora en tiempo de ciclo
    extracted_data["Mejora_Tiempo_%"] = extract_improvement_percentage(content)
    
    # Eficiencia del sistema
    extracted_data["Eficiencia_Sistema_%"] = extract_efficiency(content)
    
    # Temperaturas (buscar absorción y desorción separadamente)
    temp_abs_patterns = [
        r'temperatura[^:]{0,20}absorción[^:]{0,10}:\s*([^\n]+)',
        r'absorción[^:]{0,20}temperatura[^:]{0,10}:\s*([^\n]+)',
        r'T[^:]{0,10}absorción[^:]{0,10}:\s*([^\n]+)'
    ]
    for pattern in temp_abs_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            temp1, _ = convert_to_celsius(match.group(1))
            if temp1:
                extracted_data["Temp_Absorción_°C"] = temp1
                break
    
    temp_des_patterns = [
        r'temperatura[^:]{0,20}desorción[^:]{0,10}:\s*([^\n]+)',
        r'desorción[^:]{0,20}temperatura[^:]{0,10}:\s*([^\n]+)',
        r'T[^:]{0,10}desorción[^:]{0,10}:\s*([^\n]+)'
    ]
    for pattern in temp_des_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            temp1, _ = convert_to_celsius(match.group(1))
            if temp1:
                extracted_data["Temp_Desorción_°C"] = temp1
                break
    
    # Si no se encontraron por separado, buscar rango general
    if not extracted_data["Temp_Absorción_°C"] and extracted_data["Temperatura_Completa"]:
        temp1, temp2 = convert_to_celsius(extracted_data["Temperatura_Completa"])
        if temp1:
            extracted_data["Temp_Absorción_°C"] = temp1
        if temp2:
            extracted_data["Temp_Desorción_°C"] = temp2
    
    # Presiones (separar absorción y desorción)
    pres_abs_patterns = [
        r'presión[^:]{0,20}absorción[^:]{0,10}:\s*([^\n]+)',
        r'absorción[^:]{0,20}presión[^:]{0,10}:\s*([^\n]+)',
        r'P[^:]{0,10}absorción[^:]{0,10}:\s*([^\n]+)'
    ]
    for pattern in pres_abs_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            extracted_data["Presión_Absorción_bar"] = convert_to_bar(match.group(1))
            break
    
    pres_des_patterns = [
        r'presión[^:]{0,20}desorción[^:]{0,10}:\s*([^\n]+)',
        r'desorción[^:]{0,20}presión[^:]{0,10}:\s*([^\n]+)',
        r'P[^:]{0,10}desorción[^:]{0,10}:\s*([^\n]+)'
    ]
    for pattern in pres_des_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            extracted_data["Presión_Desorción_bar"] = convert_to_bar(match.group(1))
            break
    
    # Si no se encontraron por separado, intentar extraer de texto completo
    if not extracted_data["Presión_Absorción_bar"] and extracted_data["Presión_Completa"]:
        # Buscar dos valores en el texto completo
        bar_values = re.findall(r'([\d.]+)\s*bar', extracted_data["Presión_Completa"], re.IGNORECASE)
        if len(bar_values) >= 2:
            extracted_data["Presión_Absorción_bar"] = bar_values[0]
            extracted_data["Presión_Desorción_bar"] = bar_values[1]
        elif len(bar_values) == 1:
            extracted_data["Presión_Absorción_bar"] = bar_values[0]
    
    # Tiempos
    tiempo_abs_patterns = [
        r'tiempo[^:]{0,20}absorción[^:]{0,10}:\s*([^\n]+)',
        r'absorción[^:]{0,20}tiempo[^:]{0,10}:\s*([^\n]+)'
    ]
    for pattern in tiempo_abs_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            extracted_data["Tiempo_Absorción_min"] = convert_to_minutes(match.group(1))
            break
    
    tiempo_des_patterns = [
        r'tiempo[^:]{0,20}desorción[^:]{0,10}:\s*([^\n]+)',
        r'desorción[^:]{0,20}tiempo[^:]{0,10}:\s*([^\n]+)'
    ]
    for pattern in tiempo_des_patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            extracted_data["Tiempo_Desorción_min"] = convert_to_minutes(match.group(1))
            break
    
    # Número de tubos (ya extraído antes junto con diámetro de tubos)
    extracted_data["Número_Tubos"] = extract_number_of_tubes(content)
    
    # Información especial sobre modularidad y gestión térmica
    extracted_data["Modularidad"] = extract_modular_info(content)
    extracted_data["Sistema_Gestión_Térmica"] = extract_thermal_management(content)
    extracted_data["Resultados_Clave"] = extract_key_results(content)
    
    # Secciones específicas del proyecto
    sections_to_extract = {
        "Conclusiones_Modularidad": r"Conclusiones(?:\s+para)?(?:\s+el)?(?:\s+Diseño)?(?:\s+Modular)?[:\n]+(.*?)(?=\n##|\Z)",
        "Conclusiones_Gestión_Térmica": r"(?:Gestión|Manejo)\s+Térmica[:\n]+(.*?)(?=\n##|\Z)",
        "Aplicabilidad_Proyecto": r"(?:Aplicabilidad|Relevancia|Implicaciones)(?:\s+para)?(?:\s+el)?(?:\s+Proyecto)?[:\n]+(.*?)(?=\n##|\Z)"
    }
    
    for key, pattern in sections_to_extract.items():
        match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
        if match:
            result = match.group(1)
            extracted_data[key] = clean_value(result[:400]) + ("..." if len(result) > 400 else "")
    
    # Contar imágenes
    img_matches = len(re.findall(r"!\[.*?\]\(.*?\)", content))
    extracted_data["Imágenes_Disponibles"] = str(img_matches) if img_matches > 0 else "0"
    
    return extracted_data


def main():
    """Función principal para procesar los archivos y generar el Excel."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    all_data = []
    
    print("Iniciando extracción de datos de archivos markdown...")
    print(f"Directorio base: {base_dir}\n")
    
    md_files = list(Path(base_dir).glob("notas_*.md"))
    print(f"Archivos markdown encontrados: {len(md_files)}\n")
    
    for file in md_files:
        try:
            print(f"Procesando: {file.name}... ", end="")
            data = extract_specific_data(file)
            all_data.append(data)
            print("✓")
        except Exception as e:
            print(f"✗ Error: {str(e)}")
    
    if all_data:
        print(f"\nTotal de archivos procesados exitosamente: {len(all_data)}")
        print("Generando archivo Excel...")
        
        df = pd.DataFrame(all_data)
        df = df.reindex(columns=DEFAULT_COLUMNS)
        
        # Ordenar por año (descendente) y luego por título
        df['Año_num'] = pd.to_numeric(df['Año'], errors='coerce')
        df = df.sort_values(['Año_num', 'Título_Artículo'], ascending=[False, True])
        df = df.drop('Año_num', axis=1)
        
        output_file = os.path.join(base_dir, "matriz_informacion_articulos.xlsx")
        
        # Crear libro Excel con formato mejorado
        wb = Workbook()
        ws = wb.active
        ws.title = "Base de Conocimiento"
        
        # Configurar fila de encabezados
        ws.row_dimensions[1].height = 40
        
        for col_idx, header in enumerate(DEFAULT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Escribir datos
        for row_idx, row in enumerate(df.values, 2):
            # Configurar altura de fila para wrap text
            ws.row_dimensions[row_idx].height = 60
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = str(value) if pd.notnull(value) and value != "" else ""
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Aplicar formato especial a ciertas columnas
                if col_idx == 1:  # Archivo_Nota
                    cell.font = Font(bold=True, size=10)
                elif col_idx == 5:  # Año
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas de forma inteligente
        column_widths = {
            'Archivo_Nota': 35,
            'Referencia_PDF': 45,
            'Título_Artículo': 50,
            'Autores': 30,
            'Año': 8,
            'Revista/Fuente': 25,
            'País/Institución': 25,
            'Tipo_Estudio': 15,
            'Escala': 15,
            'Modularidad': 40,
            'Configuración_Reactor': 25,
            'Diámetro_Reactor_mm': 16,
            'Longitud_Reactor_mm': 16,
            'Volumen_Reactor_L': 14,
            'Dimensiones_Completas': 30,
            'Capacidad_H2_kg': 15,
            'Capacidad_H2_wt%': 14,
            'Capacidad_H2_Completa': 25,
            'Material_Hidruro': 20,
            'Cantidad_Hidruro_kg': 18,
            'Cantidad_Hidruro_Completa': 25,
            'Sistema_Gestión_Térmica': 40,
            'Tipo_Aletas': 25,
            'Número_Aletas': 12,
            'Espesor_Aletas_mm': 16,
            'Espaciado_Aletas_mm': 16,
            'Número_Tubos': 12,
            'Diámetro_Tubos_mm': 16,
            'Intercambiador_Calor': 25,
            'Fluido_Térmico': 15,
            'Flujo_Fluido_L_min': 16,
            'Conductividad_Térmica_W_mK': 20,
            'Mejora_Conductividad_%': 18,
            'Aditivo_Conductividad': 20,
            'Porcentaje_Aditivo_%': 18,
            'Temp_Absorción_°C': 15,
            'Temp_Desorción_°C': 15,
            'Temperatura_Completa': 30,
            'Presión_Absorción_bar': 18,
            'Presión_Desorción_bar': 18,
            'Presión_Completa': 25,
            'Tiempo_Absorción_min': 18,
            'Tiempo_Desorción_min': 18,
            'Tiempos_Completos': 30,
            'Mejora_Tiempo_%': 14,
            'Eficiencia_Sistema_%': 16,
            'Conductividad_Térmica_Mejoras': 30,
            'Resultados_Clave': 60,
            'Ventajas_Diseño': 40,
            'Limitaciones': 40,
            'Conclusiones_Modularidad': 50,
            'Conclusiones_Gestión_Térmica': 50,
            'Aplicabilidad_Proyecto': 50,
            'Imágenes_Disponibles': 10
        }
        
        for col_idx, col_name in enumerate(DEFAULT_COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_widths.get(col_name, 20)
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Añadir filtros
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(output_file)
        print(f"\n✓ Archivo Excel creado exitosamente:")
        print(f"  {output_file}")
        print(f"\nEstadísticas:")
        print(f"  - Total artículos: {len(all_data)}")
        print(f"  - Años cubiertos: {df['Año'].min()} - {df['Año'].max()}")
        print(f"  - Artículos con imágenes: {sum(1 for x in df['Imágenes_Disponibles'] if x and int(x) > 0)}")
        
    else:
        print("\n✗ No se encontraron datos para exportar.")
        print("Verifica que existan archivos notas_*.md en el directorio.")


if __name__ == "__main__":
    main()
