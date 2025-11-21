#!/usr/bin/env python3
"""
Knowledge Base Extractor v2
Extrae información técnica de artículos científicos (PDF) sobre reactores de hidruro metálico
y genera una matriz consolidada para análisis estadístico.

Basado en las categorías definidas en el informe técnico:
- Geometría del reactor (monolítico vs modular)
- Estrategias de gestión térmica (activa/pasiva)
- Escalabilidad y aplicaciones (laboratorio/industrial, móvil/estacionario)
"""

import os
import re
import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURACIÓN DE COLUMNAS PARA ANÁLISIS ESTADÍSTICO
# ============================================================================

COLUMNS = [
    # IDENTIFICACIÓN
    "ID",
    "Archivo_PDF",
    "Título_Artículo",
    "Autores",
    "Año",
    "DOI",
    "Revista",
    "País_Institución",
    
    # CLASIFICACIÓN DEL ESTUDIO
    "Tipo_Estudio",  # Experimental, Simulación, Review, Teórico
    "Escala",  # Laboratorio, Piloto, Industrial, Conceptual
    "Aplicación",  # Estacionaria, Móvil, Híbrida
    
    # GEOMETRÍA Y CONFIGURACIÓN
    "Arquitectura",  # Monolítico, Modular, Híbrido
    "Tipo_Reactor",  # Cilíndrico simple, Multi-tubular, Placas, Honeycomb, etc.
    "Cantidad_Módulos",  # Número si es modular
    "Diámetro_mm",
    "Longitud_mm",
    "Volumen_L",
    "Dimensiones_Texto",  # Descripción completa
    
    # MATERIAL Y CAPACIDAD
    "Material_Hidruro",
    "Masa_Hidruro_kg",
    "Capacidad_H2_kg",
    "Capacidad_H2_wt%",  # Gravimétrica
    "Capacidad_H2_vol",  # Volumétrica (kg H2/m³)
    
    # GESTIÓN TÉRMICA - CLASIFICACIÓN
    "Estrategia_Térmica",  # Pasiva, Activa, Híbrida
    "Método_Pasivo",  # ENG, Espuma metálica, Compactación, etc.
    "Método_Activo",  # Aletas, Tubos, PCM, etc.
    
    # GESTIÓN TÉRMICA - INTERCAMBIADORES
    "Tipo_Intercambiador",  # Camisa, Tubo central, Multi-tubular, Serpentín helicoidal
    "Configuración_Tubos",  # Cantidad y disposición
    "Diámetro_Tubos_mm",
    
    # GESTIÓN TÉRMICA - ALETAS
    "Tipo_Aletas",  # Anulares, Longitudinales, Cónicas, Disco
    "Número_Aletas",
    "Espesor_Aletas_mm",
    "Espaciado_Aletas_mm",
    "Material_Aletas",
    
    # GESTIÓN TÉRMICA - ADITIVOS
    "Aditivo_Conductividad",  # ENG, Espuma Cu/Al, Grafito
    "Porcentaje_Aditivo_%",
    "Conductividad_Base_W_mK",
    "Conductividad_Mejorada_W_mK",
    "Mejora_Conductividad_%",
    
    # CONDICIONES OPERATIVAS
    "Fluido_Térmico",
    "Caudal_L_min",
    "Temp_Absorción_°C",
    "Temp_Desorción_°C",
    "Presión_Absorción_bar",
    "Presión_Desorción_bar",
    
    # RENDIMIENTO - CINÉTICA
    "Tiempo_Absorción_min",
    "Tiempo_Desorción_min",
    "Mejora_Tiempo_vs_Base_%",
    "Referencia_Base",  # Qué diseño se usó como referencia
    
    # RENDIMIENTO - EFICIENCIA
    "Eficiencia_Carga_%",
    "Eficiencia_Descarga_%",
    "Recuperación_Térmica_%",
    "Densidad_Potencia_W_kg",
    
    # OPTIMIZACIÓN MODULAR (específico para sistemas multi-tanque)
    "Espaciado_Módulos_mm",  # Distancia S entre tanques
    "Relación_S_D",  # S/D (espaciado/diámetro)
    "Configuración_Array",  # En línea, Escalonado, etc.
    "Número_Reynolds",
    
    # ANÁLISIS Y CONCLUSIONES
    "Ventajas_Principales",
    "Limitaciones",
    "Conclusión_Modularidad",
    "Conclusión_Gestión_Térmica",
    "Conclusión_Escalabilidad",
    "Aplicabilidad_ANH951",  # Relevancia para el proyecto
    
    # DATOS ADICIONALES
    "Figuras_Relevantes",
    "Ecuaciones_Clave",
    "Notas"
]


# ============================================================================
# FUNCIONES DE EXTRACCIÓN Y PROCESAMIENTO
# ============================================================================

def extract_text_from_pdf(pdf_path):
    """
    Extrae texto de un archivo PDF usando PyPDF2.
    Retorna el texto completo del documento.
    """
    try:
        text = ""
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            num_pages = len(pdf_reader.pages)
            
            # Extraer texto de todas las páginas
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                text += page.extract_text() + "\n"
        
        return text
    except Exception as e:
        print(f"Error leyendo {pdf_path}: {e}")
        return ""


def extract_year_from_filename(filename):
    """Extrae el año del nombre del archivo."""
    match = re.search(r'(19|20)\d{2}', filename)
    return match.group(0) if match else ""


def extract_title_from_text(text):
    """
    Intenta extraer el título del artículo del texto.
    Busca en las primeras líneas del documento.
    """
    lines = text.split('\n')[:20]  # Primeras 20 líneas
    
    # El título suele estar en las primeras líneas, en mayúsculas o con formato especial
    for i, line in enumerate(lines):
        line = line.strip()
        # Saltar líneas muy cortas o típicamente no títulos
        if len(line) < 20 or len(line) > 200:
            continue
        # El título suele tener longitud moderada y no tener muchos números
        if line and not line.startswith(('http', 'www', 'doi', 'DOI')):
            return clean_text(line)
    
    return "No extraído"


def extract_doi_from_text(text):
    """Extrae el DOI del texto."""
    match = re.search(r'doi[:\s]*(10\.\d{4,}/[^\s]+)', text, re.IGNORECASE)
    return match.group(1).rstrip('.') if match else ""


def extract_authors_from_text(text):
    """Intenta extraer los autores del texto."""
    lines = text.split('\n')[:30]
    
    # Buscar patrones típicos de nombres (Apellido, Inicial.)
    for line in lines:
        if re.search(r'[A-Z][a-z]+,?\s+[A-Z]\.', line):
            # Limpiar y retornar
            authors = re.sub(r'\d+', '', line)  # Quitar números de afiliaciones
            return clean_text(authors)[:200]  # Limitar longitud
    
    return ""


def extract_reactor_type(text):
    """
    Clasifica el tipo de reactor basado en palabras clave.
    """
    text_lower = text.lower()
    
    # Patrones de búsqueda
    if any(word in text_lower for word in ['modular', 'module', 'múltiple', 'multiple', 'bank']):
        architecture = "Modular"
    elif any(word in text_lower for word in ['monolithic', 'monolítico', 'single tank', 'único']):
        architecture = "Monolítico"
    else:
        architecture = "No especificado"
    
    # Tipo específico
    if 'multi-tubular' in text_lower or 'multitubular' in text_lower:
        reactor_type = "Multi-tubular"
    elif 'plate' in text_lower or 'placa' in text_lower:
        reactor_type = "Placas"
    elif 'honeycomb' in text_lower or 'panal' in text_lower:
        reactor_type = "Honeycomb"
    elif 'shell and tube' in text_lower or 'carcasa y tubos' in text_lower:
        reactor_type = "Carcasa y tubos"
    elif 'cylindrical' in text_lower or 'cilíndrico' in text_lower:
        reactor_type = "Cilíndrico simple"
    else:
        reactor_type = "No especificado"
    
    return architecture, reactor_type


def extract_thermal_management(text):
    """
    Identifica estrategias de gestión térmica.
    """
    text_lower = text.lower()
    
    # Métodos pasivos
    passive_methods = []
    if 'eng' in text_lower or 'expanded natural graphite' in text_lower or 'grafito expandido' in text_lower:
        passive_methods.append("ENG")
    if 'foam' in text_lower or 'espuma' in text_lower:
        passive_methods.append("Espuma metálica")
    if 'compact' in text_lower and 'powder' in text_lower:
        passive_methods.append("Compactación")
    
    # Métodos activos
    active_methods = []
    if 'fin' in text_lower or 'aleta' in text_lower:
        active_methods.append("Aletas")
    if 'tube' in text_lower or 'tubo' in text_lower:
        active_methods.append("Tubos internos")
    if 'pcm' in text_lower or 'phase change material' in text_lower:
        active_methods.append("PCM")
    if 'helical' in text_lower or 'helicoidal' in text_lower or 'coil' in text_lower:
        active_methods.append("Serpentín helicoidal")
    if 'jacket' in text_lower or 'camisa' in text_lower:
        active_methods.append("Camisa")
    
    # Determinar estrategia general
    if passive_methods and active_methods:
        strategy = "Híbrida"
    elif active_methods:
        strategy = "Activa"
    elif passive_methods:
        strategy = "Pasiva"
    else:
        strategy = "No especificado"
    
    return strategy, ", ".join(passive_methods) if passive_methods else "", ", ".join(active_methods) if active_methods else ""


def extract_application_type(text):
    """
    Determina si la aplicación es estacionaria, móvil o híbrida.
    """
    text_lower = text.lower()
    
    if any(word in text_lower for word in ['stationary', 'estacionaria', 'building', 'edificio', 'grid', 'red']):
        return "Estacionaria"
    elif any(word in text_lower for word in ['mobile', 'móvil', 'vehicle', 'vehículo', 'automotive', 'automotriz']):
        return "Móvil"
    elif any(word in text_lower for word in ['portable', 'portátil']):
        return "Portátil"
    else:
        return "No especificado"


def extract_scale(text):
    """
    Determina la escala del estudio.
    """
    text_lower = text.lower()
    
    if any(word in text_lower for word in ['industrial scale', 'escala industrial', 'commercial', 'comercial']):
        return "Industrial"
    elif any(word in text_lower for word in ['pilot', 'piloto', 'prototype', 'prototipo']):
        return "Piloto"
    elif any(word in text_lower for word in ['laboratory', 'laboratorio', 'bench scale', 'lab-scale']):
        return "Laboratorio"
    elif any(word in text_lower for word in ['conceptual', 'theoretical', 'teórico']):
        return "Conceptual"
    else:
        return "No especificado"


def extract_study_type(text):
    """
    Determina el tipo de estudio.
    """
    text_lower = text.lower()
    
    if 'review' in text_lower or 'revisión' in text_lower:
        return "Review"
    elif any(word in text_lower for word in ['experimental', 'experiment']):
        return "Experimental"
    elif any(word in text_lower for word in ['numerical', 'simulation', 'cfd', 'fem', 'simulación']):
        return "Simulación"
    elif any(word in text_lower for word in ['design', 'diseño', 'optimization', 'optimización']):
        return "Diseño/Optimización"
    else:
        return "No especificado"


def extract_numeric_value(text, pattern, default=""):
    """
    Extrae un valor numérico basado en un patrón regex.
    """
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        try:
            return float(match.group(1).replace(',', '.'))
        except:
            return default
    return default


def extract_material_hydride(text):
    """
    Identifica el material de hidruro metálico usado.
    """
    text_lower = text.lower()
    
    # Lista de materiales comunes
    materials = {
        'LaNi5': ['lani5', 'lani 5'],
        'TiFe': ['tife', 'ti-fe', 'tifemn'],
        'MgH2': ['mgh2', 'mg-h2', 'magnesium hydride'],
        'NaAlH4': ['naalh4', 'sodium alanate'],
        'AB2': ['ab2', 'ab 2'],
        'AB5': ['ab5', 'ab 5'],
        'Mg2Ni': ['mg2ni', 'mg ni'],
    }
    
    for material, patterns in materials.items():
        for pattern in patterns:
            if pattern in text_lower:
                return material
    
    # Buscar patrón genérico
    match = re.search(r'([A-Z][a-z]?\d?[A-Z][a-z]?\d?)', text)
    if match:
        return match.group(1)
    
    return "No especificado"


def clean_text(text):
    """Limpia y normaliza texto extraído."""
    if not text:
        return ""
    # Eliminar múltiples espacios y saltos de línea
    text = re.sub(r'\s+', ' ', text)
    # Eliminar caracteres especiales problemáticos
    text = text.replace('\x00', '').replace('\r', '')
    return text.strip()


def process_pdf_file(pdf_path, existing_notes_path=None):
    """
    Procesa un archivo PDF y extrae información estructurada.
    Si existe un archivo de notas .md correspondiente, lo prioriza.
    """
    print(f"\nProcesando: {pdf_path.name}")
    
    # Buscar archivo de notas correspondiente
    note_text = ""
    if existing_notes_path:
        base_name = pdf_path.stem
        # Buscar archivo de notas con nombre similar
        note_files = list(existing_notes_path.glob(f"notas_*{base_name[:20]}*.md"))
        if not note_files:
            # Buscar por año
            year = extract_year_from_filename(pdf_path.name)
            if year:
                note_files = list(existing_notes_path.glob(f"notas_{year}*.md"))
        
        if note_files:
            print(f"  ✓ Archivo de notas encontrado: {note_files[0].name}")
            with open(note_files[0], 'r', encoding='utf-8') as f:
                note_text = f.read()
    
    # Extraer texto del PDF
    pdf_text = extract_text_from_pdf(pdf_path)
    
    # Combinar texto de notas (prioritario) y PDF
    combined_text = note_text + "\n" + pdf_text if note_text else pdf_text
    
    if not combined_text.strip():
        print(f"  ✗ No se pudo extraer texto")
        return None
    
    # Extraer información
    data = {
        "Archivo_PDF": pdf_path.name,
        "Año": extract_year_from_filename(pdf_path.name),
        "Título_Artículo": extract_title_from_text(combined_text),
        "DOI": extract_doi_from_text(combined_text),
        "Autores": extract_authors_from_text(combined_text),
        "Tipo_Estudio": extract_study_type(combined_text),
        "Escala": extract_scale(combined_text),
        "Aplicación": extract_application_type(combined_text),
        "Material_Hidruro": extract_material_hydride(combined_text),
    }
    
    # Arquitectura y tipo de reactor
    architecture, reactor_type = extract_reactor_type(combined_text)
    data["Arquitectura"] = architecture
    data["Tipo_Reactor"] = reactor_type
    
    # Gestión térmica
    strategy, passive, active = extract_thermal_management(combined_text)
    data["Estrategia_Térmica"] = strategy
    data["Método_Pasivo"] = passive
    data["Método_Activo"] = active
    
    print(f"  ✓ Datos extraídos: {data['Tipo_Estudio']}, {data['Escala']}, {data['Arquitectura']}")
    
    return data


def create_excel_matrix(data_list, output_path):
    """
    Crea un archivo Excel formateado con los datos extraídos.
    """
    # Crear DataFrame
    df = pd.DataFrame(data_list, columns=COLUMNS)
    
    # Agregar ID secuencial
    df['ID'] = range(1, len(df) + 1)
    
    # Guardar con formato
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Matriz_Artículos', index=False)
        
        # Obtener el workbook y worksheet
        workbook = writer.book
        worksheet = writer.sheets['Matriz_Artículos']
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Formatear encabezados
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Ajustar ancho de columna
            if column in ['Título_Artículo', 'Ventajas_Principales', 'Limitaciones']:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 40
            elif column in ['Autores', 'Material_Hidruro', 'Tipo_Reactor']:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 25
            elif column in ['Año', 'ID', 'Escala']:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 10
            else:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Formatear filas de datos
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Congelar panel superior
        worksheet.freeze_panes = 'A2'
    
    print(f"\n✓ Matriz Excel creada: {output_path}")


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main():
    """
    Función principal que coordina la extracción de información.
    """
    print("=" * 70)
    print("KNOWLEDGE BASE EXTRACTOR V2")
    print("Extracción de información técnica de artículos MH para análisis estadístico")
    print("=" * 70)
    
    # Rutas
    base_path = Path(__file__).parent.parent
    articulos_path = base_path / "articulos"
    notas_path = Path(__file__).parent
    output_path = notas_path / f"matriz_conocimiento_consolidada_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    # Verificar existencia de directorios
    if not articulos_path.exists():
        print(f"✗ Error: No se encuentra el directorio {articulos_path}")
        return
    
    print(f"\nDirectorio de artículos: {articulos_path}")
    print(f"Directorio de notas: {notas_path}")
    print(f"Archivo de salida: {output_path}")
    
    # Obtener lista de PDFs
    pdf_files = list(articulos_path.glob("*.pdf"))
    print(f"\n✓ Encontrados {len(pdf_files)} archivos PDF")
    
    if not pdf_files:
        print("✗ No se encontraron archivos PDF en el directorio")
        return
    
    # Procesar cada PDF
    data_list = []
    for pdf_file in pdf_files:
        try:
            data = process_pdf_file(pdf_file, existing_notes_path=notas_path)
            if data:
                data_list.append(data)
        except Exception as e:
            print(f"  ✗ Error procesando {pdf_file.name}: {e}")
            continue
    
    print(f"\n✓ Procesados exitosamente {len(data_list)} artículos")
    
    # Crear matriz Excel
    if data_list:
        create_excel_matrix(data_list, output_path)
        print(f"\n{'=' * 70}")
        print("PROCESO COMPLETADO")
        print(f"Total de artículos procesados: {len(data_list)}")
        print(f"Archivo generado: {output_path.name}")
        print(f"{'=' * 70}\n")
    else:
        print("\n✗ No se pudo extraer información de ningún artículo")


if __name__ == "__main__":
    main()
