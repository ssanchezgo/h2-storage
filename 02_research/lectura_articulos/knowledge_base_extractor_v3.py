#!/usr/bin/env python3
"""
Enhanced Knowledge Base Extractor V3
Versión mejorada que extrae datos cuantitativos detallados de archivos .md y PDFs
para crear una matriz completa para análisis estadístico según el informe técnico.

Categorías principales de análisis:
1. Geometría y Modularidad (monolítico vs modular)
2. Gestión Térmica (activa/pasiva, tipos de intercambiadores)
3. Rendimiento (tiempos, eficiencias, mejoras)
4. Escalabilidad (laboratorio → industrial)
"""

import os
import re
import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
import PyPDF2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# FUNCIONES DE EXTRACCIÓN DE DATOS ESTRUCTURADOS DE MARKDOWN
# ============================================================================

def extract_md_structured_data(md_path):
    """
    Extrae datos estructurados de un archivo .md de notas.
    Busca secciones específicas y campos clave.
    """
    if not md_path.exists():
        return {}
    
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except:
        return {}
    
    data = {}
    
    # Extraer título
    title_match = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
    if title_match:
        data['Título'] = title_match.group(1).strip()
    
    # Extraer autores (buscar con ** o sin ellos, también con guiones)
    authors_patterns = [
        r'\*\*Autores?:\*\*\s*([^\n]+)',
        r'[-*]\s*Autores?:\s*([^\n]+)',
        r'Autores?:\s*([^\n]+)',
    ]
    for pattern in authors_patterns:
        authors_match = re.search(pattern, content, re.MULTILINE | re.IGNORECASE)
        if authors_match:
            author_text = authors_match.group(1).strip()
            # Limpiar asteriscos residuales y espacios extra
            author_text = re.sub(r'\*+$', '', author_text).strip()
            if author_text and len(author_text) > 3:
                data['Autores'] = author_text
                break
    
    # Extraer año (buscar múltiples formatos)
    year_patterns = [
        r'\*\*Año:\*\*\s*(\d{4})',
        r'\*\*Fecha de Publicación:\*\*\s*(\d{4})',
        r'[-*]\s*Año:\s*(\d{4})',
        r'[-*]\s*Fecha.*Publicación:\s*(\d{4})',
    ]
    for pattern in year_patterns:
        year_match = re.search(pattern, content, re.IGNORECASE)
        if year_match:
            data['Año'] = year_match.group(1)
            break
    
    # Extraer revista (buscar múltiples formatos)
    revista_patterns = [
        r'\*\*Revista(?:/Fuente)?:\*\*\s*([^\n]+)',
        r'[-*]\s*Revista(?:/Fuente)?:\s*([^\n]+)',
        r'Revista(?:/Fuente)?:\s*([^\n]+)',
    ]
    for pattern in revista_patterns:
        revista_match = re.search(pattern, content, re.MULTILINE | re.IGNORECASE)
        if revista_match:
            revista_text = revista_match.group(1).strip()
            # Limpiar asteriscos residuales
            revista_text = re.sub(r'\*+$', '', revista_text).strip()
            if revista_text and len(revista_text) > 2:
                data['Revista'] = revista_text
                break
    
    # Extraer DOI (buscar múltiples formatos)
    doi_patterns = [
        r'\*\*DOI:\*\*\s*(10\.\d{4,}/[^\s\n]+)',
        r'[-*]\s*DOI:\s*(10\.\d{4,}/[^\s\n]+)',
        r'doi[:\s]+(10\.\d{4,}/[^\s\n]+)',
    ]
    for pattern in doi_patterns:
        doi_match = re.search(pattern, content, re.IGNORECASE)
        if doi_match:
            data['DOI'] = doi_match.group(1).strip()
            break
    
    # Tipo de estudio
    tipo_patterns = [
        r'\*\*Tipo de Estudio:\*\*\s*(.+?)(?:\n|$)',
        r'[-*]\s*Tipo de Estudio:\s*(.+?)(?:\n|$)',
    ]
    for pattern in tipo_patterns:
        tipo_match = re.search(pattern, content, re.MULTILINE | re.IGNORECASE)
        if tipo_match:
            data['Tipo_Estudio'] = tipo_match.group(1).strip()
            break
    
    # Escala
    escala_patterns = [
        r'\*\*Escala:\*\*\s*(.+?)(?:\n|$)',
        r'[-*]\s*Escala:\s*(.+?)(?:\n|$)',
    ]
    for pattern in escala_patterns:
        escala_match = re.search(pattern, content, re.MULTILINE | re.IGNORECASE)
        if escala_match:
            data['Escala'] = escala_match.group(1).strip()
            break
    
    # Configuración del reactor
    config_match = re.search(r'[-*]\s*Configuración.*Reactor:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    if config_match:
        data['Configuración'] = config_match.group(1).strip()
    
    # Dimensiones (buscar múltiples patrones)
    dim_patterns = [
        r'[-*]\s*Dimensiones?:\s*(.+)$',
        r'diámetro\s*(?:de|del)?\s*(?:reactor)?[:\s]*(\d+[\d.,]*)\s*(mm|cm|m)\b',
        r'longitud\s*(?:de|del)?\s*(?:reactor)?[:\s]*(\d+[\d.,]*)\s*(mm|cm|m)\b',
        r'volumen[:\s]*(\d+[\d.,]*)\s*(L|l|litros?)\b'
    ]
    
    for pattern in dim_patterns:
        match = re.search(pattern, content, re.MULTILINE | re.IGNORECASE)
        if match:
            if 'Dimensiones' not in data:
                data['Dimensiones'] = []
            data['Dimensiones'].append(match.group(0).strip())
    
    # Material de hidruro
    material_match = re.search(r'[-*]\s*Material.*Hidruro:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    if material_match:
        data['Material_Hidruro'] = material_match.group(1).strip()
    
    # Cantidad de hidruro
    cantidad_match = re.search(r'[-*]\s*Cantidad.*Hidruro:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    if cantidad_match:
        data['Cantidad_Hidruro'] = cantidad_match.group(1).strip()
    
    # Capacidad H2
    capacidad_patterns = [
        r'[-*]\s*Capacidad.*H[₂2]?:\s*(.+)$',
        r'capacidad[:\s]+(\d+[\d.,]*)\s*(kg|g|wt%|vol%)',
    ]
    
    for pattern in capacidad_patterns:
        match = re.search(pattern, content, re.MULTILINE | re.IGNORECASE)
        if match:
            if 'Capacidad_H2' not in data:
                data['Capacidad_H2'] = []
            data['Capacidad_H2'].append(match.group(0).strip())
    
    # Sistema de gestión térmica
    termica_match = re.search(r'[-*]\s*Sistema.*(?:Transferencia|Gestión).*Calor:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    if termica_match:
        data['Sistema_Térmico'] = termica_match.group(1).strip()
    
    # Tipo de aletas
    aletas_match = re.search(r'[-*]\s*Tipo.*Aletas:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    if aletas_match:
        data['Tipo_Aletas'] = aletas_match.group(1).strip()
    
    # Fluido térmico
    fluido_match = re.search(r'[-*]\s*Fluido.*[TtTérmico]:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    if fluido_match:
        data['Fluido_Térmico'] = fluido_match.group(1).strip()
    
    # Temperaturas
    temp_patterns = [
        r'[-*]\s*Temperatura.*[Oo]peración:\s*(.+)$',
        r'temperatura[:\s]+(\d+[\d.,]*)\s*[°]?[CcKk]',
        r'[-*]\s*Temperatura.*absorción:\s*(.+)$',
        r'[-*]\s*Temperatura.*desorción:\s*(.+)$',
    ]
    
    for pattern in temp_patterns:
        matches = re.findall(pattern, content, re.MULTILINE | re.IGNORECASE)
        if matches:
            if 'Temperaturas' not in data:
                data['Temperaturas'] = []
            data['Temperaturas'].extend([str(m).strip() for m in matches])
    
    # Presiones
    presion_patterns = [
        r'[-*]\s*Presión.*[Oo]peración:\s*(.+)$',
        r'presión[:\s]+(\d+[\d.,]*)\s*bar',
        r'[-*]\s*Presión.*absorción:\s*(.+)$',
        r'[-*]\s*Presión.*desorción:\s*(.+)$',
    ]
    
    for pattern in presion_patterns:
        matches = re.findall(pattern, content, re.MULTILINE | re.IGNORECASE)
        if matches:
            if 'Presiones' not in data:
                data['Presiones'] = []
            data['Presiones'].extend([str(m).strip() for m in matches])
    
    # Tiempos (absorción/desorción)
    tiempo_patterns = [
        r'[-*]\s*Tiempo.*absorción:\s*(.+)$',
        r'[-*]\s*Tiempo.*desorción:\s*(.+)$',
        r'tiempo.*carga[:\s]+(\d+[\d.,]*)\s*(min|h|s)',
        r'tiempo.*descarga[:\s]+(\d+[\d.,]*)\s*(min|h|s)',
    ]
    
    for pattern in tiempo_patterns:
        matches = re.findall(pattern, content, re.MULTILINE | re.IGNORECASE)
        if matches:
            if 'Tiempos' not in data:
                data['Tiempos'] = []
            data['Tiempos'].extend([str(m).strip() for m in matches])
    
    # Resultados clave
    resultados_section = re.search(r'##\s*Resultados.*\n(.+?)(?=##|\Z)', content, re.DOTALL | re.IGNORECASE)
    if resultados_section:
        data['Resultados'] = resultados_section.group(1).strip()[:500]  # Limitar longitud
    
    # Conclusiones sobre modularidad
    mod_section = re.search(r'###\s*Sobre Modularidad.*\n(.+?)(?=###|##|\Z)', content, re.DOTALL | re.IGNORECASE)
    if mod_section:
        data['Conclusión_Modularidad'] = mod_section.group(1).strip()[:500]
    
    # Conclusiones sobre gestión térmica
    termica_section = re.search(r'###\s*Sobre.*Gestión Térmica.*\n(.+?)(?=###|##|\Z)', content, re.DOTALL | re.IGNORECASE)
    if termica_section:
        data['Conclusión_Térmica'] = termica_section.group(1).strip()[:500]
    
    return data


# ============================================================================
# FUNCIONES DE EXTRACCIÓN NUMÉRICA Y LIMPIEZA
# ============================================================================

def extract_numeric_from_text(text, unit_pattern, convert_to_base=None):
    """
    Extrae valores numéricos con unidades específicas y convierte a unidad base.
    
    Args:
        text: Texto donde buscar
        unit_pattern: Patrón regex para la unidad (ej: 'mm|cm|m')
        convert_to_base: Función de conversión o diccionario de factores
    
    Returns:
        Lista de valores encontrados (como strings)
    """
    if not text:
        return []
    
    pattern = rf'(\d+[\d.,]*)\s*({unit_pattern})\b'
    matches = re.findall(pattern, text, re.IGNORECASE)
    
    if not matches or not convert_to_base:
        return [f"{m[0]} {m[1]}" for m in matches]
    
    # Convertir a unidad base
    converted = []
    for number, unit in matches:
        try:
            value = float(number.replace(',', '.'))
            if isinstance(convert_to_base, dict):
                factor = convert_to_base.get(unit.lower(), 1)
                converted.append(str(round(value * factor, 2)))
            else:
                converted.append(str(convert_to_base(value, unit)))
        except:
            converted.append(f"{number} {unit}")
    
    return converted


def extract_percentage_improvements(text):
    """
    Busca mejoras porcentuales reportadas en el texto.
    Ej: "mejora del 80%", "reducción de 35%", "improvement of 50%"
    """
    patterns = [
        r'(?:mejora|improvement|reducción|reduction|aumento|increase).*?(\d+[\d.,]*)\s*%',
        r'(\d+[\d.,]*)\s*%.*?(?:mejor|faster|quicker|superior)',
    ]
    
    improvements = []
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        improvements.extend([m.replace(',', '.') for m in matches])
    
    return improvements


def classify_reactor_architecture(text, config_text=""):
    """
    Clasifica la arquitectura del reactor de manera más precisa.
    """
    combined = (text + " " + config_text).lower()
    
    # Detectar modularidad
    modular_keywords = ['modular', 'multiple tanks', 'múltiples', 'bank', 'array', 'stacked']
    monolithic_keywords = ['single tank', 'único', 'monolítico', 'monolithic']
    
    is_modular = any(kw in combined for kw in modular_keywords)
    is_monolithic = any(kw in combined for kw in monolithic_keywords)
    
    if is_modular and not is_monolithic:
        return "Modular"
    elif is_monolithic and not is_modular:
        return "Monolítico"
    elif is_modular and is_monolithic:
        return "Híbrido"
    else:
        return "No especificado"


def classify_thermal_strategy(text, config_text=""):
    """
    Clasifica la estrategia de gestión térmica de forma detallada.
    """
    combined = (text + " " + config_text).lower()
    
    # Métodos pasivos
    passive = []
    if 'eng' in combined or 'expanded natural graphite' in combined:
        passive.append("ENG")
    if 'foam' in combined and ('metal' in combined or 'alumin' in combined or 'copper' in combined):
        passive.append("Espuma metálica")
    if 'compact' in combined and 'powder' in combined:
        passive.append("Compactación")
    
    # Métodos activos
    active = []
    if 'fin' in combined or 'aleta' in combined:
        # Detectar tipo de aletas
        if 'annular' in combined or 'anular' in combined:
            active.append("Aletas anulares")
        elif 'longitudinal' in combined:
            active.append("Aletas longitudinales")
        elif 'conical' in combined or 'cónica' in combined:
            active.append("Aletas cónicas")
        elif 'radial' in combined:
            active.append("Aletas radiales")
        else:
            active.append("Aletas")
    
    if 'helical' in combined or 'coil' in combined or 'spiral' in combined:
        active.append("Serpentín helicoidal")
    elif 'tube' in combined and 'multi' in combined:
        active.append("Multi-tubular")
    elif 'tube' in combined and 'central' in combined:
        active.append("Tubo central")
    
    if 'jacket' in combined or 'camisa' in combined:
        active.append("Camisa externa")
    
    if 'pcm' in combined or 'phase change' in combined:
        active.append("PCM")
    
    if 'plate' in combined and 'heat exchanger' in combined:
        active.append("Placas")
    
    # Determinar estrategia
    if passive and active:
        strategy = "Híbrida"
    elif active:
        strategy = "Activa"
    elif passive:
        strategy = "Pasiva"
    else:
        strategy = "No especificado"
    
    return strategy, ", ".join(passive), ", ".join(active)


# ============================================================================
# PROCESAMIENTO PRINCIPAL MEJORADO
# ============================================================================

def process_article_enhanced(pdf_path, notes_path):
    """
    Procesa un artículo combinando datos de .md (si existe) y PDF.
    Retorna un diccionario con todos los datos estructurados.
    """
    print(f"\n{'='*70}")
    print(f"Procesando: {pdf_path.name}")
    print(f"{'='*70}")
    
    # Inicializar diccionario de datos
    data = {
        'Archivo_PDF': pdf_path.name,
        'Año': extract_year_from_filename(pdf_path.name),
    }
    
    # 1. Buscar y cargar archivo .md correspondiente
    md_data = {}
    base_name = pdf_path.stem
    year = data['Año']
    
    # Estrategias de búsqueda de archivo .md
    search_patterns = [
        f"notas_{base_name[:30]}*.md",
        f"notas_{year}*.md" if year else "*.md",
    ]
    
    md_file = None
    for pattern in search_patterns:
        matches = list(notes_path.glob(pattern))
        if matches:
            # Elegir el que tenga nombre más similar
            md_file = max(matches, key=lambda p: len(set(p.stem.lower()) & set(base_name.lower())))
            break
    
    if md_file:
        print(f"  ✓ Archivo de notas: {md_file.name}")
        md_data = extract_md_structured_data(md_file)
        print(f"    - Campos extraídos de .md: {len(md_data)}")
    else:
        print(f"  ⚠ No se encontró archivo .md")
    
    # 2. Extraer texto del PDF
    pdf_text = ""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages[:10]:  # Primeras 10 páginas
                pdf_text += page.extract_text() + "\n"
        print(f"  ✓ Texto extraído del PDF: {len(pdf_text)} caracteres")
    except Exception as e:
        print(f"  ✗ Error leyendo PDF: {e}")
    
    # 3. Combinar y estructurar datos
    combined_text = (md_data.get('Resultados', '') + " " + 
                     md_data.get('Conclusión_Modularidad', '') + " " +
                     md_data.get('Conclusión_Térmica', '') + " " +
                     pdf_text)
    
    # Datos bibliográficos (priorizar .md)
    data['Título'] = md_data.get('Título', extract_title_from_pdf(pdf_text))
    data['Autores'] = md_data.get('Autores', extract_authors_from_pdf(pdf_text))
    data['DOI'] = md_data.get('DOI', extract_doi_from_pdf(pdf_text))
    data['Revista'] = md_data.get('Revista', '')
    
    # Clasificación del estudio
    data['Tipo_Estudio'] = md_data.get('Tipo_Estudio', classify_study_type(combined_text))
    data['Escala'] = md_data.get('Escala', classify_scale(combined_text))
    data['Aplicación'] = classify_application(combined_text)
    
    # Arquitectura y configuración
    config = md_data.get('Configuración', '')
    data['Arquitectura'] = classify_reactor_architecture(combined_text, config)
    data['Configuración_Reactor'] = config
    data['Tipo_Reactor'] = classify_reactor_type_detailed(combined_text, config)
    
    # Material de hidruro
    data['Material_Hidruro'] = md_data.get('Material_Hidruro', extract_material_hydride(combined_text))
    data['Cantidad_Hidruro'] = md_data.get('Cantidad_Hidruro', '')
    
    # Capacidad H2
    capacidad_list = md_data.get('Capacidad_H2', [])
    data['Capacidad_H2'] = "; ".join(capacidad_list) if capacidad_list else ''
    
    # Dimensiones (extraer valores numéricos)
    dim_list = md_data.get('Dimensiones', [])
    dim_text = "; ".join(dim_list) if dim_list else ''
    data['Dimensiones_Texto'] = dim_text
    
    # Extraer diámetro
    diametros = extract_numeric_from_text(
        dim_text + " " + combined_text,
        'mm|cm|m',
        {'mm': 1, 'cm': 10, 'm': 1000}
    )
    data['Diámetro_mm'] = diametros[0] if diametros else ''
    
    # Gestión térmica
    strategy, passive, active = classify_thermal_strategy(combined_text, md_data.get('Sistema_Térmico', ''))
    data['Estrategia_Térmica'] = strategy
    data['Método_Pasivo'] = passive
    data['Método_Activo'] = active
    data['Tipo_Aletas'] = md_data.get('Tipo_Aletas', '')
    data['Fluido_Térmico'] = md_data.get('Fluido_Térmico', '')
    
    # Condiciones operativas
    temps = md_data.get('Temperaturas', [])
    data['Temperaturas'] = "; ".join(temps) if temps else ''
    
    presiones = md_data.get('Presiones', [])
    data['Presiones'] = "; ".join(presiones) if presiones else ''
    
    tiempos = md_data.get('Tiempos', [])
    data['Tiempos'] = "; ".join(tiempos) if tiempos else ''
    
    # Mejoras porcentuales
    improvements = extract_percentage_improvements(combined_text)
    data['Mejoras_%'] = "; ".join(improvements) if improvements else ''
    
    # Conclusiones
    data['Conclusión_Modularidad'] = md_data.get('Conclusión_Modularidad', '')[:300]
    data['Conclusión_Térmica'] = md_data.get('Conclusión_Térmica', '')[:300]
    data['Resultados'] = md_data.get('Resultados', '')[:300]
    
    print(f"  ✓ Procesamiento completo")
    print(f"    - Arquitectura: {data['Arquitectura']}")
    print(f"    - Estrategia térmica: {data['Estrategia_Térmica']}")
    print(f"    - Material: {data['Material_Hidruro']}")
    
    return data


# Funciones auxiliares simplificadas
def extract_title_from_pdf(text):
    lines = text.split('\n')[:15]
    for line in lines:
        line = line.strip()
        if 20 < len(line) < 200 and not line.startswith(('http', 'www', 'doi')):
            return line
    return "No extraído"

def extract_authors_from_pdf(text):
    lines = text.split('\n')[:25]
    for line in lines:
        if re.search(r'[A-Z][a-z]+,?\s+[A-Z]\.', line):
            return re.sub(r'\d+', '', line).strip()[:150]
    return ""

def extract_doi_from_pdf(text):
    match = re.search(r'doi[:\s]*(10\.\d{4,}/[^\s]+)', text, re.IGNORECASE)
    return match.group(1).rstrip('.') if match else ""

def extract_year_from_filename(filename):
    match = re.search(r'(19|20)\d{2}', filename)
    return match.group(0) if match else ""

def classify_study_type(text):
    text_lower = text.lower()
    if 'review' in text_lower:
        return "Review"
    elif 'experimental' in text_lower or 'experiment' in text_lower:
        return "Experimental"
    elif 'numerical' in text_lower or 'simulation' in text_lower or 'cfd' in text_lower:
        return "Simulación"
    elif 'design' in text_lower or 'optimization' in text_lower:
        return "Diseño"
    return "No especificado"

def classify_scale(text):
    text_lower = text.lower()
    if 'industrial' in text_lower or 'commercial' in text_lower:
        return "Industrial"
    elif 'pilot' in text_lower or 'prototype' in text_lower:
        return "Piloto"
    elif 'laboratory' in text_lower or 'lab' in text_lower or 'bench' in text_lower:
        return "Laboratorio"
    return "No especificado"

def classify_application(text):
    text_lower = text.lower()
    if 'stationary' in text_lower or 'building' in text_lower:
        return "Estacionaria"
    elif 'mobile' in text_lower or 'vehicle' in text_lower or 'automotive' in text_lower:
        return "Móvil"
    return "No especificado"

def extract_material_hydride(text):
    materials = {
        'LaNi5': ['lani5'],
        'TiFe': ['tife', 'ti-fe'],
        'MgH2': ['mgh2', 'mg-h2'],
        'NaAlH4': ['naalh4'],
        'AB2': ['ab2'],
        'AB5': ['ab5'],
    }
    
    text_lower = text.lower()
    for material, patterns in materials.items():
        if any(p in text_lower for p in patterns):
            return material
    return "No especificado"

def classify_reactor_type_detailed(text, config):
    combined = (text + " " + config).lower()
    
    if 'multi-tubular' in combined or 'multitubular' in combined:
        return "Multi-tubular"
    elif 'plate' in combined:
        return "Placas"
    elif 'honeycomb' in combined:
        return "Honeycomb"
    elif 'shell and tube' in combined:
        return "Carcasa y tubos"
    elif 'cylindrical' in combined or 'cilíndrico' in combined:
        return "Cilíndrico"
    return "No especificado"


# ============================================================================
# CREACIÓN DE EXCEL MEJORADA
# ============================================================================

def create_enhanced_excel(data_list, output_path):
    """
    Crea un archivo Excel con formato profesional y múltiples hojas.
    """
    # Crear DataFrame principal
    df = pd.DataFrame(data_list)
    df.insert(0, 'ID', range(1, len(df) + 1))
    
    # Crear archivo Excel con múltiples hojas
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Hoja 1: Datos completos
        df.to_excel(writer, sheet_name='Datos_Completos', index=False)
        
        # Hoja 2: Resumen estadístico
        summary_data = {
            'Categoría': [],
            'Subcategoría': [],
            'Cantidad': []
        }
        
        # Contar por tipo de estudio
        for tipo in df['Tipo_Estudio'].value_counts().items():
            summary_data['Categoría'].append('Tipo de Estudio')
            summary_data['Subcategoría'].append(tipo[0])
            summary_data['Cantidad'].append(tipo[1])
        
        # Contar por escala
        for escala in df['Escala'].value_counts().items():
            summary_data['Categoría'].append('Escala')
            summary_data['Subcategoría'].append(escala[0])
            summary_data['Cantidad'].append(escala[1])
        
        # Contar por arquitectura
        for arq in df['Arquitectura'].value_counts().items():
            summary_data['Categoría'].append('Arquitectura')
            summary_data['Subcategoría'].append(arq[0])
            summary_data['Cantidad'].append(arq[1])
        
        # Contar por estrategia térmica
        for est in df['Estrategia_Térmica'].value_counts().items():
            summary_data['Categoría'].append('Estrategia Térmica')
            summary_data['Subcategoría'].append(est[0])
            summary_data['Cantidad'].append(est[1])
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Resumen_Estadístico', index=False)
        
        # Formatear hojas
        workbook = writer.book
        
        # Formatear hoja de datos completos
        ws_data = writer.sheets['Datos_Completos']
        format_worksheet(ws_data, df)
        
        # Formatear hoja de resumen
        ws_summary = writer.sheets['Resumen_Estadístico']
        format_worksheet(ws_summary, df_summary)
    
    print(f"\n✓ Archivo Excel creado: {output_path.name}")
    print(f"  - Hoja 1: Datos completos ({len(df)} artículos)")
    print(f"  - Hoja 2: Resumen estadístico")


def format_worksheet(ws, df):
    """Aplica formato profesional a una hoja de Excel."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Formatear encabezados
    for col_num, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Ajustar ancho
        if column in ['Título', 'Resultados', 'Conclusión_Modularidad', 'Conclusión_Térmica']:
            ws.column_dimensions[get_column_letter(col_num)].width = 50
        elif column in ['Autores', 'Configuración_Reactor']:
            ws.column_dimensions[get_column_letter(col_num)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Formatear datos
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row in range(2, len(df) + 2):
        ws.row_dimensions[row].height = 30
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Congelar panel
    ws.freeze_panes = 'B2'


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("\n" + "="*70)
    print("KNOWLEDGE BASE EXTRACTOR V3 - ENHANCED")
    print("Extracción detallada para análisis estadístico")
    print("="*70)
    
    # Rutas
    base_path = Path(__file__).parent.parent
    articulos_path = base_path / "articulos"
    notas_path = Path(__file__).parent
    output_path = notas_path / f"matriz_consolidada_v3_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    # Verificar directorios
    if not articulos_path.exists():
        print(f"\n✗ Error: No existe {articulos_path}")
        return
    
    print(f"\nRutas:")
    print(f"  - Artículos PDF: {articulos_path}")
    print(f"  - Notas .md: {notas_path}")
    print(f"  - Salida: {output_path}")
    
    # Obtener PDFs
    pdf_files = sorted(articulos_path.glob("*.pdf"))
    print(f"\n✓ Encontrados {len(pdf_files)} archivos PDF")
    
    if not pdf_files:
        print("✗ No hay PDFs para procesar")
        return
    
    # Procesar cada artículo
    data_list = []
    errors = []
    
    for i, pdf_file in enumerate(pdf_files, 1):
        print(f"\n[{i}/{len(pdf_files)}]", end=" ")
        try:
            data = process_article_enhanced(pdf_file, notas_path)
            if data:
                data_list.append(data)
        except Exception as e:
            error_msg = f"{pdf_file.name}: {str(e)}"
            errors.append(error_msg)
            print(f"  ✗ Error: {e}")
    
    # Crear Excel
    if data_list:
        print(f"\n{'='*70}")
        print("Generando archivo Excel...")
        print(f"{'='*70}")
        create_enhanced_excel(data_list, output_path)
        
        print(f"\n{'='*70}")
        print("PROCESO COMPLETADO")
        print(f"{'='*70}")
        print(f"  ✓ Artículos procesados: {len(data_list)}")
        print(f"  ✗ Errores: {len(errors)}")
        print(f"  📁 Archivo: {output_path.name}")
        print(f"{'='*70}\n")
        
        if errors:
            print("\nErrores encontrados:")
            for err in errors:
                print(f"  - {err}")
    else:
        print("\n✗ No se procesó ningún artículo exitosamente")


if __name__ == "__main__":
    main()
