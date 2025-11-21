#!/usr/bin/env python3
"""
Script para consolidar información de matriz_consolidada_v3 en la estructura de Matriz información.xlsx (Hoja Ref)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime

def extract_reference(autores, year):
    """
    Extrae una referencia en formato 'Autor et al., año'
    """
    if pd.isna(autores) or autores == 'No especificado':
        return f'Anónimo ({year})' if not pd.isna(year) else 'Anónimo'
    
    # Limpiar autores
    autores_str = str(autores).strip()
    
    # Buscar primer apellido
    # Patrón 1: "Apellido, Nombre" (separado por coma)
    if ',' in autores_str:
        parts = autores_str.split(',')
        first_author = parts[0].strip()
        # Tomar primer palabra completa (apellido)
        words = first_author.split()
        if words:
            apellido = words[0]
            # Si hay más autores (más comas)
            if autores_str.count(',') > 1 or ' and ' in autores_str.lower() or '&' in autores_str:
                return f'{apellido} et al.'
            else:
                return apellido
    
    # Patrón 2: "Nombre Apellido" (sin coma)
    # Buscar última palabra como apellido
    words = autores_str.split()
    if len(words) >= 2:
        # Filtrar palabras que parecen ser apellidos válidos
        apellido = words[-1].strip()
        # Descartar si parece copyright, reservado, etc.
        palabras_invalidas = ['reserved', 'rights', 'copyright', 'published', 'elsevier', 'ltd']
        if any(inv in apellido.lower() for inv in palabras_invalidas):
            # Intentar con primera palabra en su lugar
            apellido = words[0].strip()
        
        # Verificar si hay múltiples autores
        if ' and ' in autores_str.lower() or '&' in autores_str or ';' in autores_str:
            return f'{apellido} et al.'
        else:
            return apellido
    elif len(words) == 1:
        palabra = words[0]
        # Verificar que no sea palabra inválida
        palabras_invalidas = ['reserved', 'rights', 'copyright', 'published', 'elsevier', 'ltd', 'professor']
        if any(inv in palabra.lower() for inv in palabras_invalidas):
            return 'Anónimo'
        return palabra
    
    return 'Desconocido'


def build_description(row):
    """
    Construye descripción del sistema basada en campos disponibles
    """
    parts = []
    
    # Tipo de reactor y arquitectura
    if not pd.isna(row.get('Arquitectura')) and row.get('Arquitectura') != 'No especificado':
        parts.append(f"Arquitectura {row['Arquitectura'].lower()}")
    
    if not pd.isna(row.get('Tipo_Reactor')) and row.get('Tipo_Reactor') != 'No especificado':
        parts.append(f"tipo {row['Tipo_Reactor'].lower()}")
    
    # Dimensiones
    if not pd.isna(row.get('Dimensiones_Texto')) and row.get('Dimensiones_Texto') != 'No especificado':
        dim = str(row['Dimensiones_Texto'])
        if len(dim) < 100:  # Solo si no es muy largo
            parts.append(f"({dim})")
    
    # Estrategia térmica
    if not pd.isna(row.get('Estrategia_Térmica')) and row.get('Estrategia_Térmica') != 'No especificado':
        parts.append(f"Gestión térmica {row['Estrategia_Térmica'].lower()}")
    
    # Métodos específicos
    metodos = []
    if not pd.isna(row.get('Método_Pasivo')) and row.get('Método_Pasivo') != 'No especificado':
        metodos.append(row['Método_Pasivo'])
    if not pd.isna(row.get('Método_Activo')) and row.get('Método_Activo') != 'No especificado':
        metodos.append(row['Método_Activo'])
    
    if metodos:
        parts.append(f"con {', '.join(metodos)}")
    
    if not parts:
        # Si no hay información específica, usar título
        if not pd.isna(row.get('Título')) and row.get('Título') != 'No especificado':
            titulo = str(row['Título'])
            if len(titulo) < 150:
                return titulo
        return "Sistema de almacenamiento de hidrógeno"
    
    return '. '.join(parts)


def build_observations(row):
    """
    Construye observaciones basadas en resultados y mejoras
    """
    obs = []
    
    # Mejoras
    if not pd.isna(row.get('Mejoras_%')) and row.get('Mejoras_%') != 'No especificado':
        mejoras_str = str(row['Mejoras_%'])
        if mejoras_str and mejoras_str.strip():
            obs.append(f"Mejora: {mejoras_str}")
    
    # Tiempos
    if not pd.isna(row.get('Tiempos')) and row.get('Tiempos') != 'No especificado':
        tiempos = str(row['Tiempos'])
        if len(tiempos) < 100 and tiempos.strip():
            obs.append(f"Tiempos: {tiempos}")
    
    # Capacidad
    if not pd.isna(row.get('Capacidad_H2')) and row.get('Capacidad_H2') != 'No especificado':
        cap = str(row['Capacidad_H2'])
        if cap.strip():
            obs.append(f"Capacidad: {cap}")
    
    # Temperaturas y presiones
    if not pd.isna(row.get('Temperaturas')) and row.get('Temperaturas') != 'No especificado':
        temps = str(row['Temperaturas'])
        if len(temps) < 80 and temps.strip():
            obs.append(f"T: {temps}")
    
    if not pd.isna(row.get('Presiones')) and row.get('Presiones') != 'No especificado':
        press = str(row['Presiones'])
        if len(press) < 80 and press.strip():
            obs.append(f"P: {press}")
    
    # Conclusiones
    if not pd.isna(row.get('Conclusión_Térmica')) and row.get('Conclusión_Térmica') != 'No especificado':
        concl = str(row['Conclusión_Térmica'])
        if len(concl) < 150 and concl.strip():
            obs.append(concl)
    
    # Si no hay observaciones, buscar en Resultados
    if not obs:
        if not pd.isna(row.get('Resultados')) and row.get('Resultados') != 'No especificado':
            resultados = str(row['Resultados'])
            if len(resultados) < 200 and resultados.strip():
                obs.append(resultados)
    
    # Solo retornar vacío si realmente no hay datos
    if not obs:
        return ''
    
    return '; '.join(obs)


def consolidate_data():
    """
    Función principal de consolidación
    """
    print("=" * 80)
    print("CONSOLIDACIÓN DE DATOS EN ESTRUCTURA REF")
    print("=" * 80)
    
    # Rutas
    base_dir = Path(__file__).parent
    matriz_investigacion = base_dir.parent / '02_research' / 'lectura_articulos' / 'matriz_consolidada_v3_20251119_1515.xlsx'
    matriz_info = base_dir / 'Matriz información.xlsx'
    
    # Leer datos de investigación
    print(f"\n📖 Leyendo datos de investigación...")
    df_research = pd.read_excel(matriz_investigacion, sheet_name='Datos_Completos')
    print(f"   ✓ {len(df_research)} artículos cargados")
    
    # Crear nuevo workbook
    print(f"\n📝 Creando nueva estructura...")
    wb = openpyxl.Workbook()
    ws_ref = wb.active
    ws_ref.title = 'Ref'
    
    # Definir encabezados
    headers = ['Referencia', 'Año', 'Hidruro Metalico (MH)', 'Aplicacion', 
               'Descripción del sistema', 'Forma del Reactor', 'Observacion']
    
    for col_idx, header in enumerate(headers, 1):
        ws_ref.cell(row=1, column=col_idx, value=header)
    
    # Procesar cada artículo
    print(f"\n📝 Procesando artículos...")
    row_idx = 2  # Empezar después del encabezado
    
    for idx, row in df_research.iterrows():
        # Extraer campos
        referencia = extract_reference(row.get('Autores'), row.get('Año'))
        year = row.get('Año') if not pd.isna(row.get('Año')) else ''
        
        # Hidruro metálico
        hidruro = row.get('Material_Hidruro', 'No especificado')
        if pd.isna(hidruro) or hidruro == 'No especificado':
            hidruro = ''
        
        # Aplicación - Mejorar clasificación
        tipo_estudio = row.get('Tipo_Estudio', '')
        aplicacion_campo = row.get('Aplicación', '')
        escala = row.get('Escala', '')
        
        # Determinar tipo de estudio
        if not pd.isna(tipo_estudio) and tipo_estudio != 'No especificado':
            tipo = str(tipo_estudio)
        else:
            tipo = 'Experimental'  # Por defecto
        
        # Determinar aplicación específica
        app_especifica = ''
        if not pd.isna(aplicacion_campo) and aplicacion_campo != 'No especificado':
            app_especifica = str(aplicacion_campo)
        elif not pd.isna(escala) and escala != 'No especificado':
            escala_str = str(escala).lower()
            if 'estacionaria' in escala_str or 'industrial' in escala_str or 'laboratorio' in escala_str:
                app_especifica = 'Estacionaria'
            elif 'móvil' in escala_str or 'vehicular' in escala_str or 'transporte' in escala_str:
                app_especifica = 'Móvil'
        
        # Combinar
        if app_especifica:
            aplicacion = f'{tipo} - {app_especifica}'
        else:
            aplicacion = tipo
        
        # Descripción
        descripcion = build_description(row)
        
        # Forma del reactor
        forma = row.get('Tipo_Reactor', 'cilíndrico')
        if pd.isna(forma) or forma == 'No especificado':
            forma = 'cilíndrico'
        
        # Observaciones
        observaciones = build_observations(row)
        
        # Escribir fila
        ws_ref.cell(row=row_idx, column=1, value=referencia)
        ws_ref.cell(row=row_idx, column=2, value=year)
        ws_ref.cell(row=row_idx, column=3, value=hidruro)
        ws_ref.cell(row=row_idx, column=4, value=aplicacion)
        ws_ref.cell(row=row_idx, column=5, value=descripcion)
        ws_ref.cell(row=row_idx, column=6, value=forma)
        ws_ref.cell(row=row_idx, column=7, value=observaciones)
        
        if (idx + 1) % 10 == 0:
            print(f"   ✓ Procesados {idx + 1}/{len(df_research)} artículos...")
        
        row_idx += 1
    
    print(f"   ✓ Total procesados: {len(df_research)} artículos")
    
    # Aplicar formato
    print(f"\n🎨 Aplicando formato...")
    
    # Formato de encabezado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws_ref[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajustar anchos de columna
    ws_ref.column_dimensions['A'].width = 20  # Referencia
    ws_ref.column_dimensions['B'].width = 8   # Año
    ws_ref.column_dimensions['C'].width = 15  # Hidruro
    ws_ref.column_dimensions['D'].width = 15  # Aplicación
    ws_ref.column_dimensions['E'].width = 50  # Descripción
    ws_ref.column_dimensions['F'].width = 15  # Forma
    ws_ref.column_dimensions['G'].width = 40  # Observaciones
    
    # Aplicar bordes y alineación a todas las celdas con datos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws_ref.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Guardar
    output_path = base_dir / 'Matriz información_consolidada.xlsx'
    wb.save(output_path)
    
    print(f"\n✅ Consolidación completada!")
    print(f"   📄 Archivo: {output_path.name}")
    print(f"   📊 Artículos consolidados: {len(df_research)}")
    print("=" * 80)


if __name__ == "__main__":
    consolidate_data()
