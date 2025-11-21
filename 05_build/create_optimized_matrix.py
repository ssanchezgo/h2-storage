#!/usr/bin/env python3
"""
Script optimizado para crear matriz de información confiable
Combina y mejora consolidate_to_ref.py y create_summary_matrix.py
Para análisis estadístico robusto del proyecto ANH951
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import re


class ReferenceExtractor:
    """Extractor mejorado de referencias bibliográficas"""
    
    INVALID_WORDS = ['reserved', 'rights', 'copyright', 'published', 'elsevier', 
                     'ltd', 'professor', 'springer', 'inc', 'all']
    
    @staticmethod
    def extract(autores, year):
        """Extrae referencia en formato 'Apellido et al. (año)'"""
        if pd.isna(autores) or str(autores).strip() in ['', 'No especificado']:
            return f'Anónimo' if pd.isna(year) else f'Anónimo'
        
        autores_str = str(autores).strip()
        
        # Patrón 1: "Apellido, Nombre" (con coma)
        if ',' in autores_str:
            parts = autores_str.split(',')
            first_author = parts[0].strip()
            words = first_author.split()
            
            if words:
                apellido = words[0]
                # Validar que no sea palabra inválida
                if not any(inv in apellido.lower() for inv in ReferenceExtractor.INVALID_WORDS):
                    # Verificar múltiples autores
                    if autores_str.count(',') > 1 or ' and ' in autores_str.lower() or '&' in autores_str:
                        return f'{apellido} et al.'
                    return apellido
        
        # Patrón 2: "Nombre Apellido" (sin coma) - tomar última palabra
        words = autores_str.split()
        if len(words) >= 2:
            apellido = words[-1].strip()
            # Filtrar palabras inválidas
            if any(inv in apellido.lower() for inv in ReferenceExtractor.INVALID_WORDS):
                apellido = words[0].strip()
            
            if ' and ' in autores_str.lower() or '&' in autores_str or ';' in autores_str:
                return f'{apellido} et al.'
            return apellido
        
        elif len(words) == 1:
            palabra = words[0]
            if not any(inv in palabra.lower() for inv in ReferenceExtractor.INVALID_WORDS):
                return palabra
        
        return 'Anónimo'


class DataCleaner:
    """Limpiador y validador de datos"""
    
    @staticmethod
    def clean_text(text):
        """Limpia texto eliminando caracteres extraños"""
        if pd.isna(text) or text == 'No especificado':
            return ''
        text = str(text).strip()
        # Eliminar múltiples espacios
        text = re.sub(r'\s+', ' ', text)
        return text
    
    @staticmethod
    def extract_numeric(text):
        """Extrae primer número de un texto"""
        if pd.isna(text):
            return None
        match = re.search(r'(\d+(?:\.\d+)?)', str(text))
        return float(match.group(1)) if match else None
    
    @staticmethod
    def classify_application(row):
        """Clasifica aplicación de forma robusta"""
        tipo = DataCleaner.clean_text(row.get('Tipo_Estudio', ''))
        aplicacion = DataCleaner.clean_text(row.get('Aplicación', ''))
        escala = DataCleaner.clean_text(row.get('Escala', '')).lower()
        
        # Tipo de estudio
        if not tipo:
            tipo = 'Experimental'
        
        # Aplicación específica
        app_especifica = ''
        if aplicacion:
            app_especifica = aplicacion
        elif escala:
            if any(word in escala for word in ['estacionaria', 'industrial', 'laboratorio', 'piloto']):
                app_especifica = 'Estacionaria'
            elif any(word in escala for word in ['móvil', 'vehicular', 'transporte']):
                app_especifica = 'Móvil'
        
        return f'{tipo} - {app_especifica}' if app_especifica else tipo
    
    @staticmethod
    def normalize_reactor_type(tipo):
        """Normaliza tipo de reactor"""
        if pd.isna(tipo) or tipo == 'No especificado':
            return 'Cilíndrico'
        
        tipo_lower = str(tipo).lower()
        
        # Mapeo de variantes a nombres estándar
        mappings = {
            'cilíndrico': ['cilindrico', 'cylinder', 'tubular', 'tubo'],
            'placas': ['placa', 'plate', 'flat'],
            'multi-tubular': ['multi', 'multitubular', 'shell and tube'],
            'espiral': ['spiral', 'helicoidal', 'coil'],
            'modular': ['modular', 'module'],
            'carcasa y tubos': ['shell', 'carcasa']
        }
        
        for standard, variants in mappings.items():
            if any(var in tipo_lower for var in variants):
                return standard.title()
        
        return str(tipo).strip()
    
    @staticmethod
    def extract_capacity(text):
        """Extrae capacidad de H2 en kg"""
        if pd.isna(text):
            return None
        
        text_lower = str(text).lower()
        
        # Buscar patrones: X kg, X g
        patterns = [
            (r'(\d+(?:\.\d+)?)\s*kg', 1.0),
            (r'(\d+(?:\.\d+)?)\s*g(?:\s|$)', 0.001),
            (r'(\d+(?:\.\d+)?)\s*wt%.*?(\d+)\s*kg', lambda m: float(m.group(1)) * float(m.group(2)) / 100)
        ]
        
        for pattern, factor in patterns:
            match = re.search(pattern, text_lower)
            if match:
                if callable(factor):
                    return factor(match)
                return float(match.group(1)) * factor
        
        return None


class DescriptionBuilder:
    """Constructor de descripciones técnicas"""
    
    @staticmethod
    def build(row):
        """Construye descripción técnica del sistema"""
        parts = []
        
        # Arquitectura y tipo
        arch = DataCleaner.clean_text(row.get('Arquitectura', ''))
        tipo = DataCleaner.clean_text(row.get('Tipo_Reactor', ''))
        
        if arch:
            parts.append(f"Arquitectura {arch.lower()}")
        if tipo:
            parts.append(f"reactor tipo {DataCleaner.normalize_reactor_type(tipo).lower()}")
        
        # Estrategia térmica
        estrategia = DataCleaner.clean_text(row.get('Estrategia_Térmica', ''))
        if estrategia:
            parts.append(f"gestión térmica {estrategia.lower()}")
        
        # Métodos específicos
        metodos = []
        for col in ['Método_Pasivo', 'Método_Activo']:
            metodo = DataCleaner.clean_text(row.get(col, ''))
            if metodo:
                metodos.append(metodo)
        
        if metodos:
            parts.append(f"con {', '.join(metodos[:3])}")  # Máximo 3 métodos
        
        # Capacidad
        cap = DataCleaner.extract_capacity(row.get('Capacidad_H2', ''))
        if cap:
            parts.append(f"capacidad {cap:.2f} kg H2")
        
        if not parts:
            titulo = DataCleaner.clean_text(row.get('Título', ''))
            if titulo and len(titulo) < 150:
                return titulo
            return "Sistema de almacenamiento de hidrógeno"
        
        return '. '.join(parts).capitalize()
    
    @staticmethod
    def build_observations(row):
        """Construye observaciones técnicas"""
        obs = []
        
        # Mejoras
        mejoras = DataCleaner.clean_text(row.get('Mejoras_%', ''))
        if mejoras:
            obs.append(f"Mejora: {mejoras}")
        
        # Tiempos
        tiempos = DataCleaner.clean_text(row.get('Tiempos', ''))
        if tiempos and len(tiempos) < 80:
            obs.append(f"Tiempo: {tiempos}")
        
        # Capacidad
        cap_text = DataCleaner.clean_text(row.get('Capacidad_H2', ''))
        if cap_text and len(cap_text) < 50:
            obs.append(f"Cap: {cap_text}")
        
        # Condiciones operativas
        temps = DataCleaner.clean_text(row.get('Temperaturas', ''))
        press = DataCleaner.clean_text(row.get('Presiones', ''))
        
        if temps and len(temps) < 60:
            obs.append(f"T: {temps}")
        if press and len(press) < 60:
            obs.append(f"P: {press}")
        
        # Conclusión térmica
        concl = DataCleaner.clean_text(row.get('Conclusión_Térmica', ''))
        if concl and len(concl) < 120:
            obs.append(concl)
        
        # Si no hay nada, buscar en resultados
        if not obs:
            result = DataCleaner.clean_text(row.get('Resultados', ''))
            if result and len(result) < 150:
                return result
            return ''
        
        return '; '.join(obs)


class MatrixBuilder:
    """Constructor de matriz optimizada"""
    
    def __init__(self, source_matrix_path, output_path):
        self.source_path = Path(source_matrix_path)
        self.output_path = Path(output_path)
        self.df_source = None
        self.df_output = None
    
    def load_data(self):
        """Carga datos de la matriz de investigación"""
        print(f"\n📖 Cargando datos de: {self.source_path.name}")
        self.df_source = pd.read_excel(self.source_path, sheet_name='Datos_Completos')
        print(f"   ✓ {len(self.df_source)} artículos cargados")
        print(f"   ✓ {len(self.df_source.columns)} campos disponibles")
    
    def process_data(self):
        """Procesa y transforma los datos"""
        print(f"\n🔄 Procesando datos...")
        
        records = []
        for idx, row in self.df_source.iterrows():
            record = {
                'ID': idx + 1,
                'Referencia': ReferenceExtractor.extract(row.get('Autores'), row.get('Año')),
                'Año': row.get('Año') if not pd.isna(row.get('Año')) else '',
                'Autores_Completo': DataCleaner.clean_text(row.get('Autores', '')),
                'Titulo': DataCleaner.clean_text(row.get('Título', '')),
                'DOI': DataCleaner.clean_text(row.get('DOI', '')),
                'Hidruro_Metalico': DataCleaner.clean_text(row.get('Material_Hidruro', '')),
                'Cantidad_Hidruro_kg': DataCleaner.extract_numeric(row.get('Cantidad_Hidruro', '')),
                'Capacidad_H2_kg': DataCleaner.extract_capacity(row.get('Capacidad_H2', '')),
                'Aplicacion': DataCleaner.classify_application(row),
                'Tipo_Estudio': DataCleaner.clean_text(row.get('Tipo_Estudio', 'Experimental')),
                'Escala': DataCleaner.clean_text(row.get('Escala', '')),
                'Arquitectura': DataCleaner.clean_text(row.get('Arquitectura', '')),
                'Tipo_Reactor': DataCleaner.normalize_reactor_type(row.get('Tipo_Reactor')),
                'Descripcion_Sistema': DescriptionBuilder.build(row),
                'Estrategia_Termica': DataCleaner.clean_text(row.get('Estrategia_Térmica', '')),
                'Metodo_Pasivo': DataCleaner.clean_text(row.get('Método_Pasivo', '')),
                'Metodo_Activo': DataCleaner.clean_text(row.get('Método_Activo', '')),
                'Tipo_Aletas': DataCleaner.clean_text(row.get('Tipo_Aletas', '')),
                'Fluido_Termico': DataCleaner.clean_text(row.get('Fluido_Térmico', '')),
                'Dimensiones': DataCleaner.clean_text(row.get('Dimensiones_Texto', '')),
                'Diametro_mm': DataCleaner.extract_numeric(row.get('Diámetro_mm', '')),
                'Temperatura_Min_C': self._extract_temp(row.get('Temperaturas', ''), 'min'),
                'Temperatura_Max_C': self._extract_temp(row.get('Temperaturas', ''), 'max'),
                'Presion_Min_bar': self._extract_pressure(row.get('Presiones', ''), 'min'),
                'Presion_Max_bar': self._extract_pressure(row.get('Presiones', ''), 'max'),
                'Tiempo_Carga_min': self._extract_time(row.get('Tiempos', ''), 'carga'),
                'Tiempo_Descarga_min': self._extract_time(row.get('Tiempos', ''), 'descarga'),
                'Mejora_Porcentaje': DataCleaner.extract_numeric(row.get('Mejoras_%', '')),
                'Observaciones': DescriptionBuilder.build_observations(row),
                'Conclusion_Modularidad': DataCleaner.clean_text(row.get('Conclusión_Modularidad', '')),
                'Conclusion_Termica': DataCleaner.clean_text(row.get('Conclusión_Térmica', '')),
                'Archivo_PDF': DataCleaner.clean_text(row.get('Archivo_PDF', ''))
            }
            records.append(record)
            
            if (idx + 1) % 10 == 0:
                print(f"   ✓ Procesados {idx + 1}/{len(self.df_source)} artículos")
        
        self.df_output = pd.DataFrame(records)
        print(f"   ✓ Total procesados: {len(self.df_output)} artículos")
    
    def _extract_temp(self, text, tipo):
        """Extrae temperatura mínima o máxima"""
        if pd.isna(text):
            return None
        
        temps = re.findall(r'(\d+(?:\.\d+)?)', str(text))
        if not temps:
            return None
        
        temps_float = [float(t) for t in temps]
        return min(temps_float) if tipo == 'min' else max(temps_float)
    
    def _extract_pressure(self, text, tipo):
        """Extrae presión mínima o máxima en bar"""
        if pd.isna(text):
            return None
        
        # Buscar números con unidades
        pressures = re.findall(r'(\d+(?:\.\d+)?)\s*(bar|mpa|kpa)', str(text).lower())
        if not pressures:
            return None
        
        press_bar = []
        for val, unit in pressures:
            val_float = float(val)
            if unit == 'mpa':
                val_float *= 10
            elif unit == 'kpa':
                val_float /= 100
            press_bar.append(val_float)
        
        return min(press_bar) if tipo == 'min' else max(press_bar)
    
    def _extract_time(self, text, proceso):
        """Extrae tiempo de carga o descarga en minutos"""
        if pd.isna(text):
            return None
        
        text_lower = str(text).lower()
        
        # Buscar patrón específico
        if proceso.lower() in text_lower:
            match = re.search(rf'{proceso}.*?(\d+(?:\.\d+)?)\s*(min|h|seg)', text_lower)
            if match:
                val = float(match.group(1))
                unit = match.group(2)
                if unit == 'h':
                    val *= 60
                elif unit == 'seg':
                    val /= 60
                return val
        
        # Extraer primer tiempo encontrado
        times = re.findall(r'(\d+(?:\.\d+)?)\s*(min|h|seg)', text_lower)
        if times:
            val, unit = times[0]
            val = float(val)
            if unit == 'h':
                val *= 60
            elif unit == 'seg':
                val /= 60
            return val
        
        return None
    
    def create_excel(self):
        """Crea archivo Excel con formato"""
        print(f"\n📊 Creando archivo Excel...")
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Matriz_Datos'
        
        # Escribir encabezados
        headers = list(self.df_output.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Escribir datos
        for row_idx, row_data in enumerate(self.df_output.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Ajustar anchos
        column_widths = {
            'ID': 5, 'Referencia': 20, 'Año': 8, 'Autores_Completo': 40,
            'Titulo': 50, 'DOI': 25, 'Hidruro_Metalico': 15, 
            'Capacidad_H2_kg': 12, 'Aplicacion': 25, 'Tipo_Estudio': 15,
            'Escala': 15, 'Arquitectura': 15, 'Tipo_Reactor': 15,
            'Descripcion_Sistema': 60, 'Estrategia_Termica': 15,
            'Observaciones': 50
        }
        
        for col_idx, header in enumerate(headers, 1):
            width = column_widths.get(header, 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Aplicar bordes
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(self.df_output)+1, 
                                min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Crear hoja de resumen estadístico
        self._create_summary_sheet(wb)
        
        # Guardar
        wb.save(self.output_path)
        print(f"   ✓ Archivo guardado: {self.output_path.name}")
    
    def _create_summary_sheet(self, wb):
        """Crea hoja de resumen estadístico"""
        ws_summary = wb.create_sheet('Resumen_Estadistico')
        
        # Título
        ws_summary['A1'] = 'RESUMEN ESTADÍSTICO - BASE DE CONOCIMIENTOS ANH951'
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Estadísticas generales
        # Convertir años a numéricos para cálculos
        años_num = pd.to_numeric(self.df_output['Año'], errors='coerce')
        años_validos = años_num.dropna()
        rango_temporal = f"{años_validos.min():.0f} - {años_validos.max():.0f}" if len(años_validos) > 0 else "N/A"
        
        stats = [
            ('Total de artículos', len(self.df_output)),
            ('Rango temporal', rango_temporal),
            ('', ''),
            ('ARQUITECTURA', ''),
            *[(k, v) for k, v in self.df_output['Arquitectura'].value_counts().items()],
            ('', ''),
            ('TIPO DE REACTOR', ''),
            *[(k, v) for k, v in self.df_output['Tipo_Reactor'].value_counts().head(5).items()],
            ('', ''),
            ('HIDRUROS METÁLICOS', ''),
            *[(k, v) for k, v in self.df_output['Hidruro_Metalico'].value_counts().head(5).items()],
            ('', ''),
            ('ESTRATEGIA TÉRMICA', ''),
            *[(k, v) for k, v in self.df_output['Estrategia_Termica'].value_counts().items()],
        ]
        
        for label, value in stats:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            if label.isupper():
                ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Ajustar anchos
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 15
    
    def generate_report(self):
        """Genera reporte de calidad de datos"""
        print(f"\n📄 Generando reporte de calidad...")
        
        report_path = self.output_path.parent / 'reporte_calidad_datos.txt'
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("REPORTE DE CALIDAD DE DATOS - MATRIZ ANH951\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total de registros: {len(self.df_output)}\n\n")
            
            # Completitud de campos clave
            f.write("COMPLETITUD DE CAMPOS CLAVE:\n")
            f.write("-" * 80 + "\n")
            
            key_fields = [
                'Referencia', 'Año', 'Hidruro_Metalico', 'Capacidad_H2_kg',
                'Aplicacion', 'Arquitectura', 'Tipo_Reactor', 'Estrategia_Termica'
            ]
            
            for field in key_fields:
                if field in self.df_output.columns:
                    non_empty = self.df_output[field].notna().sum()
                    pct = (non_empty / len(self.df_output)) * 100
                    f.write(f"{field:30s}: {non_empty:3d} / {len(self.df_output):3d} ({pct:5.1f}%)\n")
            
            # Estadísticas numéricas
            f.write("\n\nESTADÍSTICAS DE CAMPOS NUMÉRICOS:\n")
            f.write("-" * 80 + "\n")
            
            numeric_fields = ['Capacidad_H2_kg', 'Diametro_mm', 'Temperatura_Max_C', 
                            'Presion_Max_bar', 'Mejora_Porcentaje']
            
            for field in numeric_fields:
                if field in self.df_output.columns:
                    data = self.df_output[field].dropna()
                    if len(data) > 0:
                        f.write(f"\n{field}:\n")
                        f.write(f"  N válidos: {len(data)}\n")
                        f.write(f"  Mínimo:    {data.min():.2f}\n")
                        f.write(f"  Máximo:    {data.max():.2f}\n")
                        f.write(f"  Media:     {data.mean():.2f}\n")
                        f.write(f"  Mediana:   {data.median():.2f}\n")
            
            # Top categorías
            f.write("\n\nTOP 5 CATEGORÍAS:\n")
            f.write("-" * 80 + "\n")
            
            for field in ['Hidruro_Metalico', 'Tipo_Reactor', 'Aplicacion']:
                if field in self.df_output.columns:
                    f.write(f"\n{field}:\n")
                    top5 = self.df_output[field].value_counts().head(5)
                    for cat, count in top5.items():
                        f.write(f"  {cat:40s}: {count:3d}\n")
        
        print(f"   ✓ Reporte guardado: {report_path.name}")
    
    def build(self):
        """Ejecuta el proceso completo"""
        print("=" * 80)
        print("CREACIÓN DE MATRIZ OPTIMIZADA PARA ANÁLISIS ESTADÍSTICO")
        print("=" * 80)
        
        self.load_data()
        self.process_data()
        self.create_excel()
        self.generate_report()
        
        print("\n" + "=" * 80)
        print("✅ PROCESO COMPLETADO")
        print("=" * 80)
        print(f"   📊 Archivo Excel: {self.output_path.name}")
        print(f"   📄 Reporte: reporte_calidad_datos.txt")
        print(f"   ✓ Registros procesados: {len(self.df_output)}")
        print(f"   ✓ Campos generados: {len(self.df_output.columns)}")
        print("=" * 80)


def main():
    """Función principal"""
    # Rutas
    base_dir = Path(__file__).parent
    source_matrix = base_dir.parent / '02_research' / 'lectura_articulos' / 'matriz_consolidada_v3_20251119_1515.xlsx'
    output_matrix = base_dir / 'Matriz_Optimizada_ANH951.xlsx'
    
    # Crear matriz
    builder = MatrixBuilder(source_matrix, output_matrix)
    builder.build()


if __name__ == "__main__":
    main()
