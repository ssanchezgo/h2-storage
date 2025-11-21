#!/usr/bin/env python3
"""
Script para crear Matriz Simplificada - Formato Hoja Ref
Estructura: 7 columnas clave para análisis estadístico
Autor: Sistema de Gestión del Conocimiento ANH951
Fecha: 2025-11-19
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

class SimplifiedMatrixBuilder:
    """Constructor de matriz simplificada formato Ref"""
    
    def __init__(self, input_file):
        self.input_file = input_file
        self.df_input = None
        self.df_output = None
        self.output_file = 'Matriz_Simplificada_ANH951.xlsx'
        
    def load_data(self):
        """Carga datos de la matriz consolidada"""
        print(f"📖 Cargando datos de: {self.input_file}")
        self.df_input = pd.read_excel(self.input_file)
        print(f"   ✓ {len(self.df_input)} artículos cargados")
        return self
    
    def extract_reference(self, autores, year):
        """Extrae referencia en formato: Apellido et al., Año"""
        if pd.isna(autores) or autores == '':
            return f"Anónimo ({year})"
        
        # Palabras a filtrar
        invalid_words = ['reserved', 'rights', 'copyright', 'published', 
                        'professor', 'elsevier', 'ltd', 'inc', 'university']
        
        autores_str = str(autores).strip()
        
        # Patrón 1: "Apellido, Nombre"
        if ',' in autores_str:
            apellido = autores_str.split(',')[0].strip()
            apellido = re.sub(r'[^\w\s-]', '', apellido)
        else:
            # Patrón 2: "Nombre Apellido"
            words = autores_str.split()
            apellido = words[-1] if words else 'Anónimo'
        
        # Limpiar caracteres especiales
        apellido = re.sub(r'[ª°\d]', '', apellido)
        apellido = apellido.strip()
        
        # Validar que no sea palabra inválida
        if apellido.lower() in invalid_words or len(apellido) < 2:
            return f"Anónimo ({year})"
        
        # Detectar múltiples autores
        if any(sep in autores_str for sep in [';', ' and ', ' y ', '&']):
            return f"{apellido} et al."
        
        return apellido
    
    def classify_application(self, row):
        """Clasifica aplicación: Tipo - Mobility"""
        # Obtener tipo de estudio (nombres con mayúscula y guión bajo)
        tipo = str(row.get('Tipo_Estudio', 'Review')).strip()
        if pd.isna(tipo) or tipo == '':
            tipo = 'Review'
        
        # Obtener aplicación directamente
        aplicacion = str(row.get('Aplicación', '')).strip()
        
        if pd.notna(aplicacion) and aplicacion != '':
            mobility = aplicacion
        else:
            # Detectar movilidad si no está especificada
            mobility_text = ' '.join([
                str(row.get('Arquitectura', '')),
                str(row.get('Tipo_Reactor', '')),
                str(row.get('Título', '')),
                str(row.get('Conclusión_Térmica', ''))
            ]).lower()
            
            # Keywords para móvil
            mobile_keywords = ['vehículo', 'transporte', 'móvil', 'automotive', 'vehicle', 
                              'car', 'onboard', 'portátil']
            
            is_mobile = any(kw in mobility_text for kw in mobile_keywords)
            mobility = "Móvil" if is_mobile else "Estacionaria"
        
        return f"{tipo} - {mobility}"
    
    def build_description(self, row):
        """Construye descripción concisa del sistema"""
        parts = []
        
        # Arquitectura
        arq = row.get('Arquitectura', '')
        if pd.notna(arq) and arq != '' and arq != 'No especificado':
            parts.append(f"Arquitectura {arq.lower()}")
        
        # Tipo reactor
        tipo = row.get('Tipo_Reactor', '')
        if pd.notna(tipo) and tipo != '' and tipo != 'No especificado':
            parts.append(f"reactor tipo {tipo.lower()}")
        
        # Dimensiones
        dim = row.get('Dimensiones_Texto', '')
        if pd.notna(dim) and dim != '' and dim != 'No especificado':
            parts.append(str(dim))
        
        # Estrategia térmica
        termica = row.get('Estrategia_Térmica', '')
        if pd.notna(termica) and termica != '' and termica != 'No especificado':
            parts.append(f"gestión térmica {termica.lower()}")
        
        # Métodos térmicos
        metodos = []
        for campo in ['Método_Pasivo', 'Método_Activo']:
            val = row.get(campo, '')
            if pd.notna(val) and val != '' and val != 'No especificado':
                metodos.append(str(val).lower())
        
        if metodos:
            parts.append(f"con {', '.join(metodos)}")
        
        description = '. '.join(parts)
        return description if description else "Sistema de almacenamiento de hidrógeno con hidruro metálico"
    
    def build_observations(self, row):
        """Construye observaciones técnicas clave"""
        obs = []
        
        # Mejoras porcentuales
        mejora = row.get('Mejoras_%', '')
        if pd.notna(mejora) and mejora != '' and mejora != 'No especificado':
            obs.append(str(mejora))
        
        # Tiempos
        tiempos = row.get('Tiempos', '')
        if pd.notna(tiempos) and tiempos != '' and tiempos != 'No especificado':
            obs.append(f"Tiempos: {tiempos}")
        
        # Capacidad
        cap = row.get('Capacidad_H2', '')
        if pd.notna(cap) and cap != '' and cap != 'No especificado':
            obs.append(f"Capacidad: {cap}")
        
        # Temperaturas
        temp = row.get('Temperaturas', '')
        if pd.notna(temp) and temp != '' and temp != 'No especificado':
            obs.append(f"T: {temp}")
        
        # Presiones
        pres = row.get('Presiones', '')
        if pd.notna(pres) and pres != '' and pres != 'No especificado':
            obs.append(f"P: {pres}")
        
        # Resultados
        result = row.get('Resultados', '')
        if pd.notna(result) and result != '' and result != 'No especificado':
            result_str = str(result)[:200]  # Limitar longitud
            obs.append(result_str)
        
        # Conclusión térmica
        concl_term = row.get('Conclusión_Térmica', '')
        if pd.notna(concl_term) and concl_term != '' and concl_term != 'No especificado':
            obs.append(str(concl_term)[:150])
        
        # Filtrar "Ver artículo original"
        observations = ', '.join(obs)
        if 'Ver artículo original' in observations or observations.strip() == '':
            return "Datos técnicos disponibles en artículo original"
        
        return observations
    
    def process_data(self):
        """Procesa datos y crea DataFrame simplificado"""
        print("\n🔄 Procesando datos...")
        
        records = []
        for idx, row in self.df_input.iterrows():
            # Extraer año
            year = row.get('Año', '')
            if pd.isna(year):
                year = 'N/A'
            else:
                year = str(year).strip()
                # Extraer solo el año numérico
                year_match = re.search(r'(19|20)\d{2}', str(year))
                if year_match:
                    year = year_match.group()
            
            record = {
                'Referencia': self.extract_reference(row.get('Autores'), year),
                'Año': year,
                'Hidruro Metalico (MH)': str(row.get('Material_Hidruro', 'No especificado')),
                'Aplicacion': self.classify_application(row),
                'Descripción del sistema': self.build_description(row),
                'Forma del Reactor': str(row.get('Tipo_Reactor', 'No especificado')),
                'Observacion': self.build_observations(row)
            }
            records.append(record)
            
            if (idx + 1) % 10 == 0:
                print(f"   ✓ Procesados {idx + 1}/{len(self.df_input)} artículos")
        
        self.df_output = pd.DataFrame(records)
        print(f"   ✓ Total procesados: {len(self.df_output)} artículos")
        return self
    
    def create_excel(self):
        """Crea archivo Excel con formato profesional"""
        print("\n📊 Creando archivo Excel...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ref"
        
        # Estilos
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Escribir encabezados
        headers = list(self.df_output.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Escribir datos
        for row_idx, row_data in enumerate(self.df_output.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = border
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 20,  # Referencia
            'B': 8,   # Año
            'C': 15,  # Hidruro Metalico
            'D': 25,  # Aplicacion
            'E': 50,  # Descripción
            'F': 18,  # Forma Reactor
            'G': 60   # Observacion
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Altura de fila del encabezado
        ws.row_dimensions[1].height = 30
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar archivo
        wb.save(self.output_file)
        print(f"   ✓ Archivo guardado: {self.output_file}")
        
        return self
    
    def generate_summary(self):
        """Genera reporte resumen"""
        print("\n📄 Generando reporte resumen...")
        
        summary = []
        summary.append("=" * 80)
        summary.append("REPORTE RESUMEN - MATRIZ SIMPLIFICADA ANH951")
        summary.append("=" * 80)
        summary.append(f"\nFecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary.append(f"Total de registros: {len(self.df_output)}")
        
        summary.append("\n\nESTRUCTURA:")
        summary.append("-" * 80)
        for i, col in enumerate(self.df_output.columns, 1):
            summary.append(f"{i}. {col}")
        
        summary.append("\n\nDISTRIBUCIÓN POR APLICACIÓN:")
        summary.append("-" * 80)
        for app, count in self.df_output['Aplicacion'].value_counts().items():
            summary.append(f"  {app:40s} : {count:3d} ({count/len(self.df_output)*100:5.1f}%)")
        
        summary.append("\n\nTOP 5 HIDRUROS METÁLICOS:")
        summary.append("-" * 80)
        for mh, count in self.df_output['Hidruro Metalico (MH)'].value_counts().head(5).items():
            summary.append(f"  {mh:40s} : {count:3d}")
        
        summary.append("\n\nDISTRIBUCIÓN TEMPORAL:")
        summary.append("-" * 80)
        años = self.df_output['Año'].value_counts().sort_index()
        summary.append(f"  Años con publicaciones: {len(años)}")
        if len(años) > 0:
            # Convertir años a numéricos para calcular rango
            años_numericos = pd.to_numeric(años.index, errors='coerce').dropna()
            if len(años_numericos) > 0:
                summary.append(f"  Rango: {int(años_numericos.min())} - {int(años_numericos.max())}")
        
        summary_text = '\n'.join(summary)
        
        # Guardar reporte
        report_file = 'reporte_matriz_simplificada.txt'
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(summary_text)
        
        print(f"   ✓ Reporte guardado: {report_file}")
        print("\n" + summary_text)
        
        return self
    
    def build(self):
        """Ejecuta el proceso completo"""
        self.load_data()
        self.process_data()
        self.create_excel()
        self.generate_summary()
        
        print("\n" + "=" * 80)
        print("✅ PROCESO COMPLETADO")
        print("=" * 80)
        print(f"   📊 Archivo: {self.output_file}")
        print(f"   ✓ Registros: {len(self.df_output)}")
        print(f"   ✓ Columnas: 7 (formato Ref)")
        print("=" * 80)

def main():
    """Función principal"""
    print("=" * 80)
    print("MATRIZ SIMPLIFICADA - FORMATO HOJA REF")
    print("Base de Conocimientos ANH951 - H2 Storage")
    print("=" * 80)
    
    # Buscar archivo de entrada más reciente
    import glob
    import os
    
    pattern = 'matriz_consolidada_v3_*.xlsx'
    files = glob.glob(pattern)
    
    if not files:
        # Buscar en directorio padre
        pattern = '../02_research/lectura_articulos/matriz_consolidada_v3_*.xlsx'
        files = glob.glob(pattern)
    
    if not files:
        print("❌ Error: No se encontró archivo matriz_consolidada_v3_*.xlsx")
        return
    
    # Usar el más reciente
    input_file = max(files, key=os.path.getmtime)
    
    # Construir matriz
    builder = SimplifiedMatrixBuilder(input_file)
    builder.build()

if __name__ == '__main__':
    main()
